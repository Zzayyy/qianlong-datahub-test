# -*- coding: utf-8 -*-
"""
性能统计模块
============
统计指标（供性能测试使用）：
  1. 每秒请求个数 / Redis 流增量（真实写入速率）
  2. 请求字节数（按秒累计）；返回字节数因数据中台未开放记为 N/A
  3. CPU 利用率（客户端进程，读 /proc，无需 psutil）
  4. SendMQ 调用延迟（µs，平均/p50/p90/p99/max）——业务响应时间因中台未开放记为 N/A
  5. 输出 Excel（按秒明细 + 汇总两个 sheet）与 log 文件

注意：本模块不依赖第三方库（openpyxl 除外，写 Excel 用）。
"""
import os
import threading
import time
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None


# ==================== CPU 采样（纯 /proc 实现）====================
def _read_total_cpu_ticks():
    """/proc/stat 第一行 cpu 的总 ticks"""
    try:
        with open("/proc/stat") as f:
            parts = f.readline().split()
        if parts and parts[0] == "cpu":
            return sum(int(x) for x in parts[1:])
    except Exception:
        pass
    return None


def _read_proc_ticks(pid):
    """/proc/<pid>/stat 的进程 CPU ticks（utime+stime+cutime+cstime）"""
    try:
        with open(f"/proc/{pid}/stat") as f:
            data = f.read()
        # 进程名可能含空格，从最后一个 ')' 后面取字段
        rp = data.rfind(")")
        parts = data[rp + 2:].split()
        # 字段 14=utime, 15=stime, 16=cutime, 17=cstime（1 起算，前移 2 后索引 11,12,13,14）
        utime = int(parts[11])
        stime = int(parts[12])
        cutime = int(parts[13])
        cstime = int(parts[14])
        return utime + stime + cutime + cstime
    except Exception:
        return None


class CpuSampler:
    """客户端进程 CPU 利用率采样。

    计算方式：进程 CPU ticks 差值 / 墙钟时间，得出"进程占用了多少个核"（可 >100%）。
    比 psutil.cpu_percent(None) 更稳（psutil 在多线程/短窗口下会返回失真值，如 6091%）。
    """

    def __init__(self):
        self.pid = os.getpid()
        self._last_proc_ticks = _read_proc_ticks(self.pid)
        self._last_wall = time.time()
        self._clock = os.sysconf("SC_CLK_TCK") if hasattr(os, "sysconf") else 100
        if not self._clock:
            self._clock = 100
        self.ncpu = os.cpu_count() or 1

    def sample(self):
        """返回进程 CPU 利用率（占用核数，%）。首次调用返回 0。"""
        now = time.time()
        proc_now = _read_proc_ticks(self.pid)
        if proc_now is None or self._last_proc_ticks is None:
            self._last_proc_ticks = proc_now
            self._last_wall = now
            return 0.0
        dt = now - self._last_wall
        if dt <= 0:
            return 0.0
        # 进程 CPU 时间（秒）= ticks / CLK_TCK
        proc_sec = (proc_now - self._last_proc_ticks) / self._clock
        # 占用核数 = proc_sec / dt，换算成百分比
        pct = proc_sec / dt * 100.0
        self._last_proc_ticks = proc_now
        self._last_wall = now
        # 合理上限：最多占满所有核（多核可 >100%）
        return min(max(pct, 0.0), 100.0 * self.ncpu)


# ==================== 性能统计 ====================
class PerfStats:
    """按秒聚合的性能统计 + 实时 CPU/Redis 采样"""

    def __init__(self, redis_cfg=None):
        """
        redis_cfg: dict(host=..., port=..., password=..., db=..., stream=...) 可选。
        提供后每秒采样一次 Redis 流的 XLEN 增量（验证真实写入）。
        """
        self.lock = threading.Lock()
        self.cpu = CpuSampler()
        self.redis_cfg = redis_cfg
        self.redis_conn = None
        self._last_xlen = None

        self.start_ts = None
        self.stop_ts = None
        self.send_start_ts = None      # 发送阶段开始（真实吞吐用）
        self.send_end_ts = None        # 发送阶段结束
        self.per_sec = OrderedDict()   # epoch秒 -> dict
        self.send_times = []           # 每次 SendMQ 耗时（µs）
        self.xadd_times = []           # 每次直写 XADD 耗时（µs，destroy 测试）
        self.send_ok = 0
        self.send_fail = 0
        self.redis_write_ok = 0        # 直写 Redis(XADD) 同步成功数（destroy 测试精确计数）
        self.bytes_ok = 0              # 成功请求的字节
        self.bytes_fail = 0            # 失败请求的字节
        self._sampler = None
        self._stop = threading.Event()

    # ---------- 发送阶段标记（用于真实吞吐，不含 wait 时间）----------
    def mark_send_start(self):
        self.send_start_ts = time.time()

    def mark_send_end(self):
        self.send_end_ts = time.time()

    # ---------- 记录 ----------
    def record_send(self, ok, payload_bytes, elapsed_us, kind="sendmq"):
        """记录一次发送：成功/字节/耗时(µs)。
        kind="sendmq" 走插件 SendMQ；kind="xadd" 为 destroy 直写 XADD（延迟单独统计）。
        注意：xadd 场景（pipeline 批量发送）传入的 elapsed_us 为批内平均延迟，非单条真实耗时。"""
        with self.lock:
            sec = int(time.time())
            rec = self.per_sec.setdefault(sec, {
                "sec": sec, "req": 0, "bytes": 0, "fail": 0,
                "cpu": 0.0, "xlen_delta": 0, "send_us_sum": 0.0, "send_us_n": 0,
                "xadd_us_sum": 0.0, "xadd_us_n": 0,
            })
            rec["req"] += 1
            rec["bytes"] += payload_bytes
            if kind == "xadd":
                rec["xadd_us_sum"] += elapsed_us
                rec["xadd_us_n"] += 1
            else:
                rec["send_us_sum"] += elapsed_us
                rec["send_us_n"] += 1
            if not ok:
                rec["fail"] += 1
                self.send_fail += 1
                self.bytes_fail += payload_bytes
            else:
                self.send_ok += 1
                self.bytes_ok += payload_bytes
            if kind == "xadd":
                self.xadd_times.append(elapsed_us)
            else:
                self.send_times.append(elapsed_us)

    def record_redis_write(self, ok):
        """记录一次直写 Redis(XADD) 的结果。

        destroy 测试走同步 XADD，成功返回即代表已真正写入流，
        可精确计数，不受采样间隔影响。
        """
        with self.lock:
            if ok:
                self.redis_write_ok += 1

    # ---------- 采样线程 ----------
    def start(self, interval=1.0):
        self.start_ts = time.time()
        self._stop.clear()
        self._sampler = threading.Thread(target=self._run, args=(interval,),
                                         daemon=True)
        self._sampler.start()

    def stop(self):
        self._stop.set()
        if self._sampler:
            self._sampler.join(timeout=3)
        self.stop_ts = time.time()
        # 先补最后一次采样（拿最终 XLEN），再关连接。
        # 注意顺序不能反：若先 close 连接，_get_redis_conn() 复用已断开对象，
        # XLEN 会抛异常返回 None，导致发送末尾的写入永远统计不到。
        self._sample_once()
        if self.redis_conn:
            try:
                self.redis_conn.close()
            except Exception:
                pass

    def _get_redis_conn(self):
        if self.redis_conn is None and self.redis_cfg:
            from mock_datahub import RespClient
            rc = self.redis_cfg
            self.redis_conn = RespClient(rc["host"], rc["port"], rc["password"],
                                         db=rc.get("db", 0))
            self.redis_conn.connect()
        return self.redis_conn

    def _get_xlen(self):
        try:
            conn = self._get_redis_conn()
            if conn is None:
                return None
            return conn.cmd("XLEN", self.redis_cfg.get("stream", "DataHub_req_stream"))
        except Exception:
            return None

    def _sample_once(self, update_xlen=True):
        cpu = self.cpu.sample()
        xlen = self._get_xlen() if update_xlen else None
        with self.lock:
            sec = int(time.time())
            rec = self.per_sec.setdefault(sec, {
                "sec": sec, "req": 0, "bytes": 0, "fail": 0,
                "cpu": 0.0, "xlen_delta": 0, "send_us_sum": 0.0, "send_us_n": 0,
                "xadd_us_sum": 0.0, "xadd_us_n": 0,
            })
            rec["cpu"] = max(rec["cpu"], cpu)   # 该秒取 CPU 峰值
            if xlen is not None:
                if self._last_xlen is not None:
                    d = max(xlen - self._last_xlen, 0)
                    # 同一秒内可能多次采样（采样线程 + stop() 补采样），增量应累加，不能覆盖
                    rec["xlen_delta"] = rec["xlen_delta"] + d
                self._last_xlen = xlen

    def _run(self, interval):
        # 高频率采 CPU（0.25s），XLEN 按传入的 interval 采样（避免 Redis 查询过频）
        last_xlen_t = 0.0
        while not self._stop.is_set():
            now = time.time()
            update_xlen = (now - last_xlen_t) >= interval
            if update_xlen:
                last_xlen_t = now
            self._sample_once(update_xlen=update_xlen)
            self._stop.wait(0.25)

    # ---------- 汇总 ----------
    def summary(self):
        """汇总指标（含分位数）"""
        with self.lock:
            total = self.send_ok + self.send_fail
            dur = (self.stop_ts or time.time()) - (self.start_ts or time.time())
            if dur <= 0:
                dur = 1e-6
            # 真实吞吐用"发送阶段"耗时（不含 wait/采样等待）
            send_dur = 0.0
            if self.send_start_ts and self.send_end_ts:
                send_dur = max(self.send_end_ts - self.send_start_ts, 1e-6)
            # SendMQ 耗时分位数（仅统计走插件的发送；destroy 直写另列 XADD 指标）
            times = sorted(self.send_times)
            xtimes = sorted(self.xadd_times)
            def pct(p, arr):
                if not arr:
                    return "N/A"
                idx = min(int(len(arr) * p), len(arr) - 1)
                return round(arr[idx], 1)
            def avg_us(arr):
                return round(sum(arr) / len(arr), 1) if arr else "N/A"
            # CPU / Redis 真实写入量（按秒采样聚合）
            # 只统计有实际请求的秒：排除启动/空闲阶段的瞬时尖峰（如插件初始化占满多核）
            active = [rec for rec in self.per_sec.values() if rec.get("req", 0) > 0]
            cpu_vals = [rec.get("cpu", 0) for rec in active if rec.get("cpu", 0) > 0]
            # XLEN 采样增量（参考值）：对全部秒求和，避免最后一次补采样落在空闲秒被漏掉
            redis_inc = sum(rec.get("xlen_delta", 0) for rec in self.per_sec.values())
            # 有直写 XADD 精确计数则优先使用（destroy 场景与"总请求数"对齐），否则退回采样值
            redis_written = self.redis_write_ok if self.redis_write_ok else redis_inc
            return {
                "总请求数": total,
                "成功数": self.send_ok,
                "失败数": self.send_fail,
                "成功率%": (self.send_ok / total * 100) if total else 0.0,
                "总耗时(含等待,s)": round(dur, 3),
                "发送耗时(s)": round(send_dur, 3) if send_dur else "N/A",
                "吞吐(条/s,按发送耗时)": round(total / send_dur, 2) if send_dur else 0.0,
                "请求总字节(B)": self.bytes_ok + self.bytes_fail,
                "请求字节(KB/s,按发送耗时)": round((self.bytes_ok + self.bytes_fail) / send_dur / 1024, 2) if send_dur else 0.0,
                "平均单请求字节(B)": round((self.bytes_ok + self.bytes_fail) / total, 1) if total else 0,
                "SendMQ平均(µs)": avg_us(times),
                "SendMQ p50(µs)": pct(0.50, times),
                "SendMQ p90(µs)": pct(0.90, times),
                "SendMQ p99(µs)": pct(0.99, times),
                "SendMQ max(µs)": round(times[-1], 1) if times else "N/A",
                # destroy 直写 XADD 延迟（pipeline 批量发送时为批内平均，非单条真实耗时）
                "XADD均(µs)": avg_us(xtimes),
                "XADD p50(µs)": pct(0.50, xtimes),
                "XADD p90(µs)": pct(0.90, xtimes),
                "XADD p99(µs)": pct(0.99, xtimes),
                "XADD max(µs)": round(xtimes[-1], 1) if xtimes else "N/A",
                "CPU平均%": round(sum(cpu_vals) / len(cpu_vals), 1) if cpu_vals else 0.0,
                "CPU峰值%": round(max(cpu_vals), 1) if cpu_vals else 0.0,
                "Redis写入增量": redis_written,
                "Redis写入(采样参考)": redis_inc,
                "每秒采样点数": len(self.per_sec),
                # 数据中台未开放
                "返回字节(B)": "N/A(数据中台未开放)",
                "业务响应时间(µs)": "N/A(数据中台未开放)",
            }

    # ---------- 输出 ----------
    def detail_text(self):
        """按秒明细文本（供运行日志末尾追加）"""
        lines = ["\n" + "=" * 60, "性能统计(按秒明细)", "=" * 60]
        lines.append("  秒 | 请求数 | 字节数 | 失败 | CPU% | Redis流增量 | SendMQ均µs | XADD均µs")
        with self.lock:
            for sec, rec in self.per_sec.items():
                avg = (rec["send_us_sum"] / rec["send_us_n"]) if rec.get("send_us_n") else 0
                xavg = (rec["xadd_us_sum"] / rec["xadd_us_n"]) if rec.get("xadd_us_n") else 0
                lines.append(f"  {rec['sec']} | {rec['req']} | {rec['bytes']} | {rec['fail']} | "
                             f"{rec['cpu']:.1f} | {rec['xlen_delta']} | {avg:.1f} | {xavg:.1f}")
        return "\n".join(lines)

    def save_log(self, path):
        s = self.summary()
        lines = [
            "=" * 60,
            "性能测试统计",
            "=" * 60,
        ]
        for k, v in s.items():
            lines.append(f"{k}: {v}")
        lines.append("=" * 60)
        lines.append("按秒明细:")
        lines.append("  秒 | 请求数 | 字节数 | 失败 | CPU% | Redis流增量 | SendMQ均µs | XADD均µs")
        with self.lock:
            for sec, rec in self.per_sec.items():
                avg = (rec["send_us_sum"] / rec["send_us_n"]) if rec.get("send_us_n") else 0
                xavg = (rec["xadd_us_sum"] / rec["xadd_us_n"]) if rec.get("xadd_us_n") else 0
                lines.append(f"  {rec['sec']} | {rec['req']} | {rec['bytes']} | {rec['fail']} | "
                             f"{rec['cpu']:.1f} | {rec['xlen_delta']} | {avg:.1f} | {xavg:.1f}")
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return path

    def save_excel(self, path):
        """保存为 Excel：Sheet1 按秒明细，Sheet2 汇总"""
        if Workbook is None:
            return None
        s = self.summary()
        wb = Workbook()
        # Sheet1 按秒明细
        ws = wb.active
        ws.title = "按秒统计"
        bold = Font(bold=True)
        headers = ["秒", "请求数", "字节数(B)", "失败数", "CPU%", "Redis流增量",
                   "SendMQ平均(µs)", "XADD平均(µs)"]
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        with self.lock:
            for sec, rec in self.per_sec.items():
                avg = (rec["send_us_sum"] / rec["send_us_n"]) if rec.get("send_us_n") else 0
                xavg = (rec["xadd_us_sum"] / rec["xadd_us_n"]) if rec.get("xadd_us_n") else 0
                ws.append([rec["sec"], rec["req"], rec["bytes"], rec["fail"],
                           round(rec["cpu"], 1), rec["xlen_delta"], round(avg, 1),
                           round(xavg, 1)])
        # Sheet2 汇总
        ws2 = wb.create_sheet("汇总")
        ws2.append(["指标", "值"])
        for c in ws2[1]:
            c.font = bold
        for k, v in s.items():
            ws2.append([k, v])
        # 列宽
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 28
        wb.save(path)
        return path
