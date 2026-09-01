# -*- coding: utf-8 -*-
"""
多线程压测 GUI（PySide6）
=========================
封装 make_excel.py / send_test.py 两个命令行工具：
  - 选择接口 -> 生成 Excel（本地）
  - 配置并发参数 -> 运行发送测试
      * 未启用远程：本地运行（只生成报文，.so 在 Linux 时才真正发送）
      * 启用远程：通过 SSH 在 Linux 上执行，自动上传数据文件并实时回传日志

依赖：venv 环境已装 PySide6 + paramiko（pip install PySide6 paramiko）
运行：venv/Scripts/python.exe gui_test.py
"""
import json
import os
import subprocess
import sys
import time

# 屏蔽 Qt 在 Windows 上枚举旧系统字体失败的无害警告（Fixedsys/MS Sans Serif 等）
os.environ.setdefault("QT_LOGGING_RULES", "qt.qpa.fonts.warning=false")

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QTextCursor
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QComboBox, QPushButton, QSpinBox,
    QDoubleSpinBox, QCheckBox, QLineEdit, QTextEdit, QGroupBox,
    QGridLayout, QVBoxLayout, QHBoxLayout, QMessageBox,
    QListWidget, QListWidgetItem, QAbstractItemView,
    QTableWidget, QTableWidgetItem, QHeaderView, QSplitter, QFileDialog,
    QTabWidget, QToolButton, QScrollArea, QFrame, QSizePolicy,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
INTERFACES_DIR = os.path.join(BASE_DIR, "interfaces")
PYTHON = sys.executable

# 日志区最大行数：超出后丢弃最旧的行，防止日志过多导致界面卡死
MAX_LOG_LINES = 3000


def list_interfaces():
    """扫描 interfaces/ 下的接口定义文件"""
    if not os.path.isdir(INTERFACES_DIR):
        return []
    return sorted(f[:-3] for f in os.listdir(INTERFACES_DIR)
                  if f.endswith(".py") and not f.startswith(("__", "_")))


# ==================== 本地 Worker ====================
class Worker(QThread):
    """本地子进程执行（支持多条命令循环），实时回传输出"""
    line = Signal(str)
    finished = Signal(int)

    def __init__(self, cmds, cwd, parent=None):
        super().__init__(parent)
        # cmds: 单条 [..] 或多条 [[..],[..]]，依次执行
        if cmds and isinstance(cmds[0], list):
            self.cmds = cmds
        else:
            self.cmds = [cmds]
        self.cwd = cwd
        self._proc = None

    def run(self):
        rc_last = 0
        try:
            env = dict(os.environ)
            env["PYTHONIOENCODING"] = "utf-8"
            for cmd in self.cmds:
                self.line.emit("$ " + " ".join(cmd))
                self._proc = subprocess.Popen(
                    cmd, cwd=self.cwd, stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT, text=True, encoding="utf-8",
                    errors="replace", bufsize=1, env=env)
                for out in self._proc.stdout:
                    self.line.emit(out.rstrip("\n"))
                self._proc.wait()
                rc_last = self._proc.returncode
                if rc_last != 0:
                    break
            self.finished.emit(rc_last)
        except Exception as e:
            self.line.emit(f"[ERROR] {e}")
            self.finished.emit(-1)

    def stop(self):
        if self._proc and self._proc.poll() is None:
            try:
                self._proc.terminate()
            except Exception:
                pass


# ==================== SSH Worker ====================
class SshWorker(QThread):
    """通过 SSH 在远程 Linux 上执行命令，实时回传输出"""
    line = Signal(str)
    finished = Signal(int)

    def __init__(self, host, port, username, password, remote_dir, command,
                 files_to_upload=None, download_config=None, parent=None,
                 ini_update=None):
        super().__init__(parent)
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.remote_dir = remote_dir
        self.command = command
        self.files_to_upload = files_to_upload or []
        # download_config: {"remote": 远程目录, "local": 本地目录, "patterns": [glob...]}
        self.download_config = download_config
        # ini_update: {"kv": {REDISHOST:..., REDISPORT:...}, } 更新远程 DataHub.ini [REDIS] 段
        self.ini_update = ini_update
        self._client = None
        self._chan = None
        self._result_snap = {}   # 执行前的远程结果目录快照 {文件名: mtime}

    def run(self):
        try:
            import paramiko
        except ImportError:
            self.line.emit("[ERROR] 缺少 paramiko，请先: pip install paramiko")
            self.finished.emit(-1)
            return
        try:
            self._client = paramiko.SSHClient()
            self._client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self._client.connect(self.host, port=self.port,
                                 username=self.username, password=self.password,
                                 timeout=15)
            self.line.emit(f"[SSH] 已连接 {self.host}")

            # 更新远程 DataHub.ini [REDIS] 段（保留其他段落）
            if self.ini_update:
                sftp0 = self._client.open_sftp()
                try:
                    ini_path = f"{self.remote_dir}/DataHub.ini"
                    text = ""
                    try:
                        with sftp0.open(ini_path, "r") as fh:
                            text = fh.read().decode("utf-8", "replace")
                    except IOError:
                        text = "[REDIS]\n"
                    import re
                    for key, val in self.ini_update["kv"].items():
                        pat = re.compile(rf"^{key}.*$", re.M)
                        if pat.search(text):
                            text = pat.sub(f"{key}={val}", text)
                        else:
                            text = text.rstrip("\n") + f"\n{key}={val}\n"
                    with sftp0.open(ini_path, "w") as fh:
                        fh.write(text.encode("utf-8"))
                    for key, val in self.ini_update["kv"].items():
                        self.line.emit(f"[SSH] DataHub.ini {key}={val}")
                finally:
                    sftp0.close()

            # 上传文件（数据文件 + 脚本，保持最新）
            if self.files_to_upload:
                sftp = self._client.open_sftp()

                def ensure_dir(path):
                    """递归确保远程目录存在"""
                    parts = path.split("/")
                    cur = ""
                    for p in parts:
                        cur = f"{cur}/{p}".replace("//", "/")
                        if not cur or p == "":
                            continue
                        try:
                            sftp.stat(cur)
                        except IOError:
                            try:
                                sftp.mkdir(cur)
                            except IOError:
                                pass

                # 确保远程根目录 + 子目录存在
                ensure_dir(self.remote_dir)

                def needs_upload(local, remote):
                    """变更检测：远程不存在，或本地比远程新，则需上传"""
                    if not os.path.exists(local):
                        return False
                    try:
                        rstat = sftp.stat(remote)
                    except IOError:
                        return True   # 远程不存在
                    local_mtime = os.path.getmtime(local)
                    remote_mtime = rstat.st_mtime
                    return local_mtime > remote_mtime

                for local, remote in self.files_to_upload:
                    dirpath = "/".join(remote.split("/")[:-1])
                    ensure_dir(dirpath)
                    if not needs_upload(local, remote):
                        self.line.emit(f"[SSH] 跳过上传（已是最新） {os.path.basename(local)}")
                        continue
                    try:
                        sftp.put(local, remote)
                        self.line.emit(f"[SSH] 已上传 {os.path.basename(local)}")
                    except Exception as e:
                        self.line.emit(f"[SSH] 上传 {local} 失败: {e}")
                sftp.close()

            # 远程执行
            full_cmd = f"cd {self.remote_dir} && {self.command}"
            self.line.emit(f"$ {full_cmd}")
            # 记录结果目录快照：执行后只下载本次新增/更新的文件，避免全量下载历史版本
            self._result_snap = self._snapshot_results() if self.download_config else {}
            self._chan = self._client.get_transport().open_session()
            self._chan.settimeout(0)
            self._chan.exec_command(full_cmd)
            bufs = {"out": "", "err": ""}
            while True:
                # stdout
                if self._chan.recv_ready():
                    data = self._chan.recv(4096).decode("utf-8", "replace")
                    self._drain("out", bufs, data)
                # stderr（import 错误等都会到 stderr，不读就看不到）
                if self._chan.recv_stderr_ready():
                    data = self._chan.recv_stderr(4096).decode("utf-8", "replace")
                    self._drain("err", bufs, data)
                if self._chan.exit_status_ready() and not self._chan.recv_ready() \
                        and not self._chan.recv_stderr_ready():
                    # 收尾
                    while self._chan.recv_ready():
                        self._drain("out", bufs, self._chan.recv(4096).decode("utf-8", "replace"))
                    while self._chan.recv_stderr_ready():
                        self._drain("err", bufs, self._chan.recv_stderr(4096).decode("utf-8", "replace"))
                    for name in ("out", "err"):
                        if bufs[name].strip():
                            for ln in bufs[name].split("\n"):
                                if ln.strip():
                                    self.line.emit(ln.rstrip("\r"))
                    rc = self._chan.recv_exit_status()
                    # 下载远程结果文件（性能统计 Excel/log）
                    if self.download_config:
                        try:
                            self._download_results(rc)
                        except Exception as e:
                            self.line.emit(f"[SSH] 下载结果失败: {e}")
                    self.finished.emit(rc)
                    return
                self.msleep(50)
        except Exception as e:
            self.line.emit(f"[SSH][ERROR] {e}")
            self.finished.emit(-1)
        finally:
            if self._client:
                try:
                    self._client.close()
                except Exception:
                    pass

    def _snapshot_results(self):
        """记录远程结果目录现有文件 {remote#文件名: mtime}，用于区分本次新增"""
        try:
            sftp = self._client.open_sftp()
            try:
                snap = {}
                for d in self.download_config.get("dirs", []):
                    try:
                        for attr in sftp.listdir_attr(d["remote"]):
                            snap[f"{d['remote']}#{attr.filename}"] = attr.st_mtime
                    except IOError:
                        pass
                return snap
            finally:
                sftp.close()
        except Exception:
            return {}

    def _download_results(self, rc):
        """执行完后，按目录下载本次新增/更新的结果文件（stats JSON/xlsx/运行日志）"""
        cfg = self.download_config
        import fnmatch
        sftp = self._client.open_sftp()
        try:
            downloaded = 0
            skipped = 0
            for d in cfg.get("dirs", []):
                remote_dir = d["remote"]
                local_dir = d["local"]
                patterns = d.get("patterns", [])
                os.makedirs(local_dir, exist_ok=True)
                try:
                    files = sftp.listdir_attr(remote_dir)
                except IOError:
                    self.line.emit(f"[SSH] 远程目录 {remote_dir} 不存在，跳过")
                    continue
                for attr in sorted(files, key=lambda a: a.st_mtime, reverse=True):
                    name = attr.filename
                    if not any(fnmatch.fnmatch(name, p) for p in patterns):
                        continue
                    old_mtime = self._result_snap.get(f"{remote_dir}#{name}")
                    if old_mtime is not None and attr.st_mtime <= old_mtime + 0.001:
                        skipped += 1      # 历史文件，本次未更新
                        continue
                    local_path = os.path.join(local_dir, name)
                    try:
                        sftp.get(f"{remote_dir}/{name}", local_path)
                        self.line.emit(f"[SSH] 已下载结果: {local_path}")
                        downloaded += 1
                    except Exception as e:
                        self.line.emit(f"[SSH] 下载 {name} 失败: {e}")
            if downloaded == 0:
                if skipped:
                    self.line.emit(f"[SSH] 远程无本次新增结果文件（跳过历史文件 {skipped} 个）")
                else:
                    self.line.emit("[SSH] 远程无匹配结果文件（未生成统计/日志）")
            else:
                self.line.emit(f"[SSH] 结果已下载"
                               + (f"（跳过历史文件 {skipped} 个）" if skipped else ""))
        finally:
            sftp.close()

    def _drain(self, name, bufs, data):
        """按行消费缓冲，超过一定量先强制刷出，避免越积越多"""
        bufs[name] += data
        buf = bufs[name]
        while "\n" in buf:
            line, buf = buf.split("\n", 1)
            if line.strip():
                self.line.emit(line.rstrip("\r"))
        bufs[name] = buf

    def stop(self):
        try:
            if self._chan:
                self._chan.close()
            if self._client:
                self._client.close()
        except Exception:
            pass


# ==================== 批量汇总加载线程 ====================
class SummaryLoader(QThread):
    """后台扫描下载目录 *_stats.json 并解析（磁盘 IO 不阻塞 GUI）"""
    done = Signal(object, int, str)   # (rows, file_count, error)

    def __init__(self, download_dir, batch_start, parent=None):
        super().__init__(parent)
        self.download_dir = download_dir
        self.batch_start = batch_start

    def run(self):
        try:
            files = []
            if os.path.isdir(self.download_dir):
                for fn in os.listdir(self.download_dir):
                    if fn.endswith("_stats.json"):
                        full = os.path.join(self.download_dir, fn)
                        try:
                            with open(full, encoding="utf-8") as f:
                                data = json.load(f)
                        except Exception:
                            continue
                        try:
                            files.append((os.path.getmtime(full), full, data))
                        except Exception:
                            continue
            if self.batch_start:
                # 只统计本次批量开始后下载的文件
                files = [x for x in files if x[0] >= self.batch_start - 5]
            file_count = len(files)
            by_name = {}
            for _mtime, full, data in sorted(files, key=lambda x: x[0]):
                if isinstance(data, dict):
                    by_name[data.get("interface") or os.path.basename(full)] = data
            rows = [(n, d["summary"]) for n, d in by_name.items()
                    if isinstance(d, dict) and isinstance(d.get("summary"), dict)]
            self.done.emit(rows, file_count, "")
        except Exception as e:
            self.done.emit([], 0, str(e))


# ==================== 配置 ====================
CONFIG_PATH = os.path.join(BASE_DIR, "config.ini")

DEFAULT_CONFIG = {
    "host": "192.168.1.136",
    "port": "22",
    "username": "yangsh",
    "password": "qianlong@135246",
    "remote_dir": "/home/yangsh/so_test",
    "workers": "4",
    "max": "0",
    "wait": "5.0",
    "mock": "1",
    "destroy_via_plugin": "0",
    "destroy_mode": "type1",
    "download": "1",
    "download_dir": "out/performance",
    "remote": "1",       # 启用远程执行
    "quiet": "1",        # 安静模式
    "box_redis": "1",    # 右侧三个标题条的展开状态
    "box_linux": "1",
    "box_out": "1",
    "hsplit": "",        # 上半左右分栏比例（自动记忆）
    "vsplit": "",        # 上下分栏比例（自动记忆）
}

# 批量汇总表格列：(表头, summary JSON 里的键)。interface 取顶层字段
SUMMARY_COLS = [
    ("接口", "interface"),
    ("总请求数", "总请求数"),
    ("成功", "成功数"),
    ("失败", "失败数"),
    ("成功率%", "成功率%"),
    ("吞吐(条/s)", "吞吐(条/s,按发送耗时)"),
    ("CPU平均%", "CPU平均%"),
    ("CPU峰值%", "CPU峰值%"),
    ("Redis写入", "Redis写入增量"),
    ("发送耗时(s)", "发送耗时(s)"),
    ("总耗时(s)", "总耗时(含等待,s)"),
    ("请求KB/s", "请求字节(KB/s,按发送耗时)"),
    ("SendMQ均(µs)", "SendMQ平均(µs)"),
    ("p50(µs)", "SendMQ p50(µs)"),
    ("p90(µs)", "SendMQ p90(µs)"),
    ("p99(µs)", "SendMQ p99(µs)"),
    ("max(µs)", "SendMQ max(µs)"),
    ("XADD均(µs)", "XADD均(µs)"),
    ("XADD p99(µs)", "XADD p99(µs)"),
]


def load_config():
    """从 config.ini 读取配置；文件不存在/缺失项用默认值"""
    cfg = dict(DEFAULT_CONFIG)
    try:
        import configparser
        cp = configparser.ConfigParser()
        cp.read(CONFIG_PATH, encoding="utf-8")
        if cp.has_section("main"):
            for k in cfg:
                if cp.has_option("main", k):
                    cfg[k] = cp.get("main", k)
    except Exception:
        pass
    return cfg


def save_config(cfg):
    """保存配置到 config.ini"""
    try:
        import configparser
        cp = configparser.ConfigParser()
        cp.add_section("main")
        for k, v in cfg.items():
            cp.set("main", k, str(v))
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            cp.write(f)
    except Exception:
        pass


# ==================== 可折叠分组框 ====================
class CollapsibleBox(QWidget):
    """标题条 + 内容：标题样式沿用 QToolBox 页签风格，默认展开，点击可收起"""

    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.btn = QToolButton()
        self.btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.btn.setArrowType(Qt.ArrowType.DownArrow)
        self.btn.setText(title)
        self.btn.setCheckable(True)
        self.btn.setChecked(True)
        self.btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        f = QFont("Microsoft YaHei", 9)
        f.setBold(True)
        self.btn.setFont(f)
        # 与 QToolBox::tab 保持一致的视觉：浅蓝底 + 圆角 + 加粗，hover 为选中色
        self.btn.setStyleSheet(
            "QToolButton { background: #eef3fb; border: 1px solid #d0d8e4;"
            " border-radius: 3px; padding: 5px 10px; font-weight: bold;"
            " text-align: left; }"
            "QToolButton:hover { background: #cfe0f8; }")

        self.content = QWidget()
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(2)
        lay.addWidget(self.btn)
        lay.addWidget(self.content)
        self.btn.clicked.connect(self._on_toggled)

    def _on_toggled(self, checked):
        self.btn.setArrowType(Qt.ArrowType.DownArrow if checked
                              else Qt.ArrowType.RightArrow)
        self.content.setVisible(checked)

    def setExpanded(self, expanded):
        self.btn.setChecked(expanded)
        self._on_toggled(expanded)

    def isExpanded(self):
        return self.btn.isChecked()


# ==================== 主窗口 ====================
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("数据中台压力测试工具 (DataHub 压测客户端)")
        self.resize(1380, 900)
        self.worker = None
        self.cfg = load_config()
        self._batch_names = []
        self._batch_idx = 0
        self._batch_stopped = False
        self._batch_start = 0        # 本次批量开始时间戳（汇总过滤用）
        self._auto_export = False    # 批量完成后自动导出汇总总表
        self._batch_params = {}      # 本次批量参数快照（导出到汇总 Excel 的"批次信息"sheet）
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(6)

        # ---- 顶栏：标题居中 + 下方连接摘要（不展开配置也能确认）----
        title_lbl = QLabel("DataHub 数据中台 · 条件单压测工具")
        tfont = QFont("Microsoft YaHei", 13)
        tfont.setBold(True)
        title_lbl.setFont(tfont)
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root.addWidget(title_lbl)

        conn_row = QHBoxLayout()
        conn_row.setContentsMargins(0, 0, 0, 0)
        conn_row.addStretch(1)
        self.lbl_conn = QLabel("")
        self.lbl_conn.setStyleSheet("color: #666;")
        self.lbl_conn.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_conn.setToolTip("当前生效的连接配置，可在右侧「连接设置」中修改")
        conn_row.addWidget(self.lbl_conn)
        conn_row.addStretch(1)
        root.addLayout(conn_row)

        # ---- 主体：上=主流程|连接设置，下=日志/汇总（比例可拖拽并自动记忆）----
        self.vsplit = QSplitter(Qt.Orientation.Vertical)
        self.vsplit.setChildrenCollapsible(False)
        self.hsplit = QSplitter(Qt.Orientation.Horizontal)
        self.hsplit.setChildrenCollapsible(False)

        # ============ 左上：主流程（1 测试数据 -> 2 发送参数 -> 运行）============
        left_w = QWidget()
        left_lay = QVBoxLayout(left_w)
        left_lay.setContentsMargins(0, 0, 0, 0)

        # ---- 1. 测试数据（接口复选框平铺）----
        grp1 = QGroupBox("1. 测试数据 (勾选要发送的接口)")
        g1 = QGridLayout(grp1)
        names = list_interfaces()
        self.chk_ifaces = {}     # name -> QCheckBox
        cols = 4
        for i, name in enumerate(names):
            chk = QCheckBox(name)
            chk.setChecked(True)              # 默认全选
            chk.toggled.connect(self.update_interfaces_label)
            self.chk_ifaces[name] = chk
            g1.addWidget(chk, i // cols, i % cols)
        for c in range(cols):
            g1.setColumnStretch(c, 1)

        # 右侧：已选计数 + 全选/清空/生成
        right = QVBoxLayout()
        self.lbl_excel = QLabel("")
        slfont = QFont("Microsoft YaHei", 10)
        slfont.setBold(True)
        self.lbl_excel.setFont(slfont)
        right.addWidget(self.lbl_excel)
        right.addStretch(1)
        self.btn_sel_all = QPushButton("全选")
        self.btn_sel_all.clicked.connect(lambda: self._set_all_iface(True))
        right.addWidget(self.btn_sel_all)
        self.btn_sel_none = QPushButton("清空")
        self.btn_sel_none.clicked.connect(lambda: self._set_all_iface(False))
        right.addWidget(self.btn_sel_none)
        self.btn_make = QPushButton("批量生成 Excel")
        self.btn_make.clicked.connect(self.on_make)
        right.addWidget(self.btn_make)
        rows = max((len(names) + cols - 1) // cols, 1)
        g1.addLayout(right, 0, cols, rows, 1)
        left_lay.addWidget(grp1)

        # ============ 右上：连接设置（标题条样式不变，内容默认全部展开）============
        self.box_redis = CollapsibleBox("Redis 配置")
        self.box_linux = CollapsibleBox("远程 Linux")
        self.box_out = CollapsibleBox("结果保存")
        self.box_redis.setExpanded(self.cfg.get("box_redis", "1") == "1")
        self.box_linux.setExpanded(self.cfg.get("box_linux", "1") == "1")
        self.box_out.setExpanded(self.cfg.get("box_out", "1") == "1")
        for b in (self.box_redis, self.box_linux, self.box_out):
            b.btn.clicked.connect(self._save_ui_config)   # 折叠状态立即记住

        # ---- Redis 配置（写入远程 DataHub.ini 的 [REDIS] 段）----
        gr = QGridLayout(self.box_redis.content)
        gr.addWidget(QLabel("主机:"), 0, 0)
        self.edit_r_host = QLineEdit(self.cfg.get("r_host", "192.168.1.137"))
        gr.addWidget(self.edit_r_host, 0, 1)
        gr.addWidget(QLabel("端口:"), 0, 2)
        self.edit_r_port = QLineEdit(self.cfg.get("r_port", "6379"))
        gr.addWidget(self.edit_r_port, 0, 3)
        gr.addWidget(QLabel("密码:"), 1, 0)
        self.edit_r_pwd = QLineEdit(self.cfg.get("r_pwd", "QianLong@2026&"))
        self.edit_r_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        gr.addWidget(self.edit_r_pwd, 1, 1)
        gr.addWidget(QLabel("数据库(SELECT):"), 1, 2)
        self.spin_r_db = QSpinBox()
        self.spin_r_db.setRange(0, 15)
        self.spin_r_db.setValue(int(self.cfg.get("r_db", "2")))
        gr.addWidget(self.spin_r_db, 1, 3)
        self.btn_read_redis = QPushButton("读取远程配置")
        self.btn_read_redis.clicked.connect(self.on_read_redis)
        gr.addWidget(self.btn_read_redis, 0, 4)
        self.btn_save_redis = QPushButton("保存到远程")
        self.btn_save_redis.clicked.connect(self.on_save_redis)
        gr.addWidget(self.btn_save_redis, 1, 4)

        # ---- 2. 发送参数 ----
        grp2 = QGroupBox("2. 发送参数")
        g2 = QGridLayout(grp2)
        # 输入列自动拉伸，让两列输入框宽度均匀分配
        g2.setColumnStretch(1, 1)
        g2.setColumnStretch(3, 1)

        # 行0：压测规模（并发 / 条数）
        g2.addWidget(QLabel("并发线程数:"), 0, 0)
        self.spin_workers = QSpinBox()
        self.spin_workers.setRange(1, 500)
        self.spin_workers.setValue(int(self.cfg.get("workers", "4")))
        g2.addWidget(self.spin_workers, 0, 1)

        g2.addWidget(QLabel("最多条数(0=全部):"), 0, 2)
        self.spin_max = QSpinBox()
        self.spin_max.setRange(0, 10000000)
        self.spin_max.setValue(int(self.cfg.get("max", "0")))
        g2.addWidget(self.spin_max, 0, 3)

        # 行1：等待回复 / 安静模式
        g2.addWidget(QLabel("等待回复秒数:"), 1, 0)
        self.spin_wait = QDoubleSpinBox()
        self.spin_wait.setRange(0, 300)
        self.spin_wait.setValue(float(self.cfg.get("wait", "5.0")))
        g2.addWidget(self.spin_wait, 1, 1)

        self.chk_quiet = QCheckBox("安静模式 (批量压测建议勾选，减少日志)")
        self.chk_quiet.setChecked(self.cfg.get("quiet", "1") == "1")
        g2.addWidget(self.chk_quiet, 1, 2, 1, 2)

        # 行2：发送模式开关
        self.chk_mock = QCheckBox("模拟数据中台应答器 (mock)")
        self.chk_mock.setChecked(self.cfg.get("mock", "1") == "1")
        g2.addWidget(self.chk_mock, 2, 0, 1, 2)

        self.chk_destroy_plugin = QCheckBox("破坏数据走插件 (默认直写 Redis)")
        self.chk_destroy_plugin.setChecked(self.cfg.get("destroy_via_plugin", "0") == "1")
        g2.addWidget(self.chk_destroy_plugin, 2, 2, 1, 2)

        # 行3：用例类型过滤（--type）
        g2.addWidget(QLabel("用例类型:"), 3, 0)
        type_box = QHBoxLayout()
        self.chk_type_normal = QCheckBox("normal")
        self.chk_type_error = QCheckBox("error")
        self.chk_type_destroy = QCheckBox("destroy")
        self.chk_type_normal.setChecked(True)
        self.chk_type_error.setChecked(True)
        self.chk_type_destroy.setChecked(True)
        for cb in (self.chk_type_normal, self.chk_type_error, self.chk_type_destroy):
            type_box.addWidget(cb)
        type_box.addStretch(1)
        g2.addLayout(type_box, 3, 1, 1, 3)

        # 行4：破坏测试类型（直写 Redis 时生效；mixed=交替 type1/type2）
        g2.addWidget(QLabel("破坏类型:"), 4, 0)
        self.combo_destroy_mode = QComboBox()
        self.combo_destroy_mode.addItem("type1 核心字段正常 / 业务数据畸形", "type1")
        self.combo_destroy_mode.addItem("type2 核心字段乱填", "type2")
        self.combo_destroy_mode.addItem("mixed 两种交替", "mixed")
        _dm_idx = self.combo_destroy_mode.findData(self.cfg.get("destroy_mode", "type1"))
        self.combo_destroy_mode.setCurrentIndex(_dm_idx if _dm_idx >= 0 else 0)
        self.combo_destroy_mode.setToolTip(
            "type1：外层核心字段正常，只把业务数据(task)写成畸形\n"
            "type2：外层核心字段(来源/回执信息)乱填\n"
            "mixed：两种交替各一半")
        g2.addWidget(self.combo_destroy_mode, 4, 1, 1, 3)
        left_lay.addWidget(grp2)

        # ---- 远程 Linux（发送测试在其上执行）----
        g3 = QGridLayout(self.box_linux.content)
        self.chk_remote = QCheckBox("启用远程执行（发送测试在其上执行）")
        self.chk_remote.setChecked(self.cfg.get("remote", "1") == "1")
        g3.addWidget(self.chk_remote, 0, 0, 1, 4)

        g3.addWidget(QLabel("主机:"), 1, 0)
        self.edit_host = QLineEdit(self.cfg.get("host", "192.168.1.136"))
        g3.addWidget(self.edit_host, 1, 1)
        g3.addWidget(QLabel("端口:"), 1, 2)
        self.spin_ssh_port = QSpinBox()
        self.spin_ssh_port.setRange(1, 65535)
        self.spin_ssh_port.setValue(int(self.cfg.get("port", "22")))
        g3.addWidget(self.spin_ssh_port, 1, 3)

        g3.addWidget(QLabel("用户名:"), 2, 0)
        self.edit_user = QLineEdit(self.cfg.get("username", "yangsh"))
        g3.addWidget(self.edit_user, 2, 1)
        g3.addWidget(QLabel("密码:"), 2, 2)
        self.edit_pass = QLineEdit(self.cfg.get("password", ""))
        self.edit_pass.setEchoMode(QLineEdit.EchoMode.Password)
        g3.addWidget(self.edit_pass, 2, 3)

        g3.addWidget(QLabel("远程目录:"), 3, 0)
        self.edit_remote_dir = QLineEdit(self.cfg.get("remote_dir", "/home/yangsh/so_test"))
        g3.addWidget(self.edit_remote_dir, 3, 1, 1, 3)

        # ---- 结果保存（性能统计 Excel/log）----
        g4 = QGridLayout(self.box_out.content)
        self.chk_download = QCheckBox("自动下载结果到本地电脑")
        self.chk_download.setChecked(self.cfg.get("download", "1") == "1")
        g4.addWidget(self.chk_download, 0, 0, 1, 4)
        g4.addWidget(QLabel("本地目录:"), 1, 0)
        _dl_dir = self.cfg.get("download_dir", "out/performance")
        if not os.path.isabs(_dl_dir):
            _dl_dir = os.path.join(BASE_DIR, _dl_dir)
        self.edit_download_dir = QLineEdit(_dl_dir)
        g4.addWidget(self.edit_download_dir, 1, 1, 1, 2)
        self.btn_pick_dir = QPushButton("选择目录…")
        self.btn_pick_dir.setToolTip("选择性能统计结果的本地保存目录（会自动记住）")
        self.btn_pick_dir.clicked.connect(self.on_pick_download_dir)
        g4.addWidget(self.btn_pick_dir, 1, 3)

        # ---- 上半组装：左=主流程，右=连接设置 + 使用提示（可滚动，窗口矮也不挤）----
        right_w = QScrollArea()
        right_w.setWidgetResizable(True)
        right_w.setFrameShape(QFrame.Shape.NoFrame)
        inner = QWidget()
        right_lay = QVBoxLayout(inner)
        right_lay.setContentsMargins(0, 0, 4, 0)
        right_lay.addWidget(self.box_redis)
        right_lay.addWidget(self.box_linux)
        right_lay.addWidget(self.box_out)

        tip = QGroupBox("使用提示")
        tlay = QVBoxLayout(tip)
        tip_text = QLabel(
            "· 不勾选「启用远程执行」时只在本机生成报文，"
            "真实发送需要远程 Linux（.so 在 Linux 上）<br>"
            "· Redis / 远程 Linux 改完自动生效并记住，无需另外保存<br>"
            "· 标题栏可点击收起，分栏比例可拖拽，下次启动保持"
        )
        tip_text.setWordWrap(True)
        tip_text.setTextFormat(Qt.TextFormat.RichText)
        tip_text.setStyleSheet("color: #555;")
        tip_text.setAlignment(Qt.AlignmentFlag.AlignTop)
        tlay.addWidget(tip_text)
        right_lay.addWidget(tip)
        right_lay.addStretch(1)
        right_w.setWidget(inner)

        self.hsplit.addWidget(left_w)
        self.hsplit.addWidget(right_w)
        self.hsplit.setStretchFactor(0, 3)
        self.hsplit.setStretchFactor(1, 2)

        # ---- 操作按钮（紧跟主流程）----
        btns = QHBoxLayout()
        self.btn_send = QPushButton("运行发送测试")
        self.btn_send.clicked.connect(self.on_send)
        self.btn_stop = QPushButton("停止")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.on_stop)
        self.btn_clear = QPushButton("清空日志")
        self.btn_clear.clicked.connect(lambda: self.log.clear())
        btns.addWidget(self.btn_send)
        btns.addWidget(self.btn_stop)
        btns.addWidget(self.btn_clear)
        self.btn_upload_all = QPushButton("批量上传所有接口数据")
        self.btn_upload_all.clicked.connect(self.on_upload_all)
        btns.addWidget(self.btn_upload_all)
        btns.addStretch(1)
        left_lay.addLayout(btns)
        left_lay.addStretch(1)

        # ========== 下半：日志 / 汇总（全宽，汇总表 19 列不再需要横向滚动）==========
        self.tabs = QTabWidget()

        # ---- 页1：运行日志 ----
        self.tab_log = QWidget()
        tl = QVBoxLayout(self.tab_log)
        tl.setContentsMargins(4, 4, 4, 4)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        # 文档最大块数：Qt 内部高效丢弃最旧行，防止日志过多导致渲染卡死
        self.log.document().setMaximumBlockCount(MAX_LOG_LINES)
        font = QFont("Consolas", 9)
        self.log.setFont(font)
        tl.addWidget(self.log)
        self.tabs.addTab(self.tab_log, "运行日志")

        # ---- 页2：批量汇总分析 ----
        self.tab_summary = QWidget()
        ts = QVBoxLayout(self.tab_summary)
        ts.setContentsMargins(4, 6, 4, 4)
        top5 = QHBoxLayout()
        self.lbl_summary_info = QLabel("批量发送完成后自动汇总")
        self.lbl_summary_info.setStyleSheet("color: #666;")
        top5.addWidget(self.lbl_summary_info)
        top5.addStretch(1)
        self.btn_summary_refresh = QPushButton("刷新汇总")
        self.btn_summary_refresh.clicked.connect(self._refresh_summary)
        top5.addWidget(self.btn_summary_refresh)
        self.btn_summary_export = QPushButton("导出 Excel")
        self.btn_summary_export.clicked.connect(self.on_export_summary)
        top5.addWidget(self.btn_summary_export)
        self.btn_summary_clear = QPushButton("清空")
        self.btn_summary_clear.clicked.connect(self.on_clear_summary)
        top5.addWidget(self.btn_summary_clear)
        ts.addLayout(top5)
        self.table_summary = QTableWidget(0, len(SUMMARY_COLS))
        self.table_summary.setHorizontalHeaderLabels([c[0] for c in SUMMARY_COLS])
        self.table_summary.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_summary.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_summary.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_summary.verticalHeader().setVisible(False)
        # 全宽后 19 列基本一屏可见：接口列拉伸，其余固定宽，仍超出时才横向滚动
        hdr = self.table_summary.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setDefaultSectionSize(84)
        hdr.setMinimumSectionSize(60)
        hdr.resizeSection(0, 130)
        self.table_summary.setStyleSheet(
            "QTableWidget { background: #fafafa; border: 1px solid #ddd; border-radius: 4px; }"
            "QHeaderView::section { background: #e8f0fe; font-weight: bold;"
            " border: none; border-right: 1px solid #d0d0d0; padding: 4px; }"
            "QTableWidget::item { padding: 2px 6px; }")
        ts.addWidget(self.table_summary, 1)
        self.tabs.addTab(self.tab_summary, "批量汇总分析")

        # ---- 上下组装 ----
        self.vsplit.addWidget(self.hsplit)
        self.vsplit.addWidget(self.tabs)
        self.vsplit.setStretchFactor(0, 0)
        self.vsplit.setStretchFactor(1, 1)
        root.addWidget(self.vsplit, 1)

        # ---- 顶栏连接摘要随配置实时刷新 ----
        for w in (self.edit_r_host, self.edit_r_port, self.edit_host,
                  self.edit_user, self.edit_remote_dir):
            w.textChanged.connect(self._update_conn_summary)
        for sp in (self.spin_r_db, self.spin_ssh_port):
            sp.valueChanged.connect(self._update_conn_summary)
        self.chk_remote.toggled.connect(self._update_conn_summary)

        self._restore_splitter("hsplit", self.hsplit, [880, 500])
        self._restore_splitter("vsplit", self.vsplit, [480, 362])
        self.update_excel_label()
        self._update_conn_summary()

    # ---------- 布局辅助 ----------
    def _update_conn_summary(self):
        """顶栏摘要：不用展开“连接设置”也能确认当前连的是哪台机器"""
        r = (f"{self.edit_r_host.text().strip()}:{self.edit_r_port.text().strip()}"
             f"/{self.spin_r_db.value()}")
        if self.chk_remote.isChecked():
            exe = (f"{self.edit_user.text().strip()}@"
                   f"{self.edit_host.text().strip()}:{self.spin_ssh_port.value()}")
        else:
            exe = "本地执行"
        self.lbl_conn.setText(f"Redis {r}    |    执行 {exe}")

    def _restore_splitter(self, key, splitter, default):
        """恢复上次的分割比例；配置无效则用默认值"""
        raw = self.cfg.get(key, "")
        try:
            sizes = [int(x) for x in raw.split(",") if x.strip()]
            if len(sizes) == len(default) and all(s > 0 for s in sizes):
                splitter.setSizes(sizes)
                return
        except Exception:
            pass
        splitter.setSizes(list(default))

    # ---------- 工具 ----------
    def selected_interfaces(self):
        """返回勾选的接口名列表"""
        return [n for n, chk in self.chk_ifaces.items() if chk.isChecked()]

    def _set_all_iface(self, checked):
        for chk in self.chk_ifaces.values():
            chk.setChecked(checked)

    def update_interfaces_label(self):
        sel = self.selected_interfaces()
        total = len(self.chk_ifaces)
        if sel:
            text = f"已选 {len(sel)}/{total}"
        else:
            text = "未选择任何接口"
        self.lbl_excel.setText(text)

    def update_excel_label(self):
        self.update_interfaces_label()

    # ---------- 批量汇总 ----------
    def _download_dir_abs(self):
        d = self.edit_download_dir.text().strip()
        if not os.path.isabs(d):
            d = os.path.join(BASE_DIR, d)
        return d

    def on_pick_download_dir(self):
        """弹窗选择本地结果目录（不存在时询问是否创建），选择后立即记住"""
        cur = self._download_dir_abs()
        start = cur if os.path.isdir(cur) else BASE_DIR
        d = QFileDialog.getExistingDirectory(
            self, "选择结果保存目录", start,
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontResolveSymlinks)
        if not d:
            return
        d = os.path.normpath(d)
        if not os.path.isdir(d):
            if QMessageBox.question(
                    self, "目录不存在", f"目录不存在，是否创建？\n{d}",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) != QMessageBox.StandardButton.Yes:
                return
            try:
                os.makedirs(d, exist_ok=True)
            except Exception as e:
                QMessageBox.warning(self, "提示", f"创建目录失败: {e}")
                return
        self.edit_download_dir.setText(d)
        cfg = load_config()
        cfg["download_dir"] = d
        save_config(cfg)
        self.append_log(f"[配置] 结果保存目录已设为 {d}")

    def _safe_float(self, v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _refresh_summary(self):
        """后台扫描 stats JSON 并填表（磁盘 IO 不阻塞 UI）。

        仅当本次会话运行过发送测试（_batch_start 非 0）才去扫描下载目录，
        避免显示历史批次的数据。
        """
        if not self._batch_start:
            # 本次会话还没发过测试：无数据可汇总
            self._fill_summary_table([])
            self.lbl_summary_info.setText("尚未运行发送测试：请先点击“运行发送测试”，批量结束后自动汇总")
            return
        loader = getattr(self, "_summary_loader", None)
        if loader and loader.isRunning():
            self.append_log("[汇总] 正在刷新中，请稍候...")
            return
        loader = SummaryLoader(self._download_dir_abs(), self._batch_start, self)
        self._summary_loader = loader
        self.btn_summary_refresh.setEnabled(False)
        self.lbl_summary_info.setText("正在扫描统计文件...")
        loader.done.connect(self._on_summary_loaded)
        loader.start()

    def _on_summary_loaded(self, rows, file_count, error):
        """汇总线程完成回调（GUI 主线程）"""
        if hasattr(self, "btn_summary_refresh"):
            self.btn_summary_refresh.setEnabled(True)
        if error:
            self.append_log(f"[汇总] 刷新失败: {error}")
            self.lbl_summary_info.setText(f"刷新失败: {error}")
            return
        self._fill_summary_table(rows)
        if rows:
            self.lbl_summary_info.setText(
                f"汇总 {len(rows)} 个接口（{file_count} 个统计文件，本次批量）")
            # 批量完成后自动导出汇总总表（仅一次）
            if self._auto_export:
                self._auto_export = False
                path = self._do_export_summary(self._collect_summary_rows(),
                                               self._batch_params)
                if path:
                    self.append_log(f"[EXPORT] 批量完成，自动导出汇总总表: {path}")
                else:
                    self.append_log("[WARN] 汇总总表自动导出失败（缺 openpyxl）")
        else:
            self.lbl_summary_info.setText(
                "暂无本次批量统计文件（远程未生成 *_stats.json？）")

    def _fill_summary_table(self, rows):
        """rows: [(接口名, summary dict), ...]。列数固定，禁止自动扩列"""
        table = self.table_summary
        table.setColumnCount(len(SUMMARY_COLS))     # 固定列数，杜绝 setItem 越界扩列
        table.setRowCount(len(rows) + (1 if rows else 0))
        for r, (name, s) in enumerate(rows):
            vals = [name] + [s.get(c[1], "") for c in SUMMARY_COLS[1:]]
            self._set_summary_row(r, vals, bold=False)
        if rows:
            n = len(SUMMARY_COLS)
            agg = [""] * n
            total = sum(self._safe_float(s.get("总请求数")) for _, s in rows)
            ok = sum(self._safe_float(s.get("成功数")) for _, s in rows)
            fail = sum(self._safe_float(s.get("失败数")) for _, s in rows)
            thr = sum(self._safe_float(s.get("吞吐(条/s,按发送耗时)")) for _, s in rows)
            cpu_avg = [self._safe_float(s.get("CPU平均%")) for _, s in rows]
            cpu_avg = [v for v in cpu_avg if v > 0]
            cpu_peak = [self._safe_float(s.get("CPU峰值%")) for _, s in rows]
            cpu_peak = [v for v in cpu_peak if v > 0]
            redis_inc = sum(self._safe_float(s.get("Redis写入增量")) for _, s in rows)
            agg[0] = "合计"
            agg[1] = total
            agg[2] = ok
            agg[3] = fail
            agg[4] = round(ok / total * 100, 2) if total else 0
            agg[5] = round(thr, 2)
            agg[6] = round(sum(cpu_avg) / len(cpu_avg), 1) if cpu_avg else ""
            agg[7] = round(max(cpu_peak), 1) if cpu_peak else ""
            agg[8] = redis_inc
            self._set_summary_row(len(rows), agg, bold=True)

    def _set_summary_row(self, row, vals, bold=False):
        table = self.table_summary
        for c, v in enumerate(vals):
            if c >= table.columnCount():
                break
            item = QTableWidgetItem(str(v))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if bold:
                f = QFont("Microsoft YaHei", 9)
                f.setBold(True)
                item.setFont(f)
                item.setBackground(Qt.GlobalColor.lightGray)
            table.setItem(row, c, item)

    def on_clear_summary(self):
        self.table_summary.setRowCount(0)
        self.lbl_summary_info.setText("已清空汇总显示（下次批量后自动刷新）")

    def _do_export_summary(self, rows, params=None):
        """把汇总行数据导出为 Excel，返回路径（不弹窗；失败返回 None）。
        params: 批次参数 dict，写入"批次信息"sheet"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return None
        wb = Workbook()
        ws = wb.active
        ws.title = "批量汇总"
        ws.append([c[0] for c in SUMMARY_COLS])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E8F0FE")
        for r in rows:
            ws.append(r)
        if params:
            ws2 = wb.create_sheet("批次信息")
            ws2.append(["参数", "值"])
            for c in ws2[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="E8F0FE")
            for k, v in params.items():
                ws2.append([k, v])
            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 60
        ts = time.strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self._download_dir_abs(), f"批量汇总_{ts}.xlsx")
        wb.save(path)
        return path

    def on_export_summary(self):
        """把汇总表导出为 Excel（所见即所得，含合计行）"""
        rows = self._collect_summary_rows()
        if not rows:
            QMessageBox.warning(self, "提示", "当前没有汇总数据，请先批量发送并刷新")
            return
        path = self._do_export_summary(rows, self._batch_params)
        if not path:
            QMessageBox.warning(self, "提示", "缺少 openpyxl，请先: pip install openpyxl")
            return
        self.append_log(f"[EXPORT] 已导出批量汇总: {path}")
        QMessageBox.information(self, "完成", f"已导出:\n{path}")

    def _collect_summary_rows(self):
        """从表格控件读取当前所有行数据（含合计）"""
        table = self.table_summary
        rows = []
        for r in range(table.rowCount()):
            vals = []
            for c in range(table.columnCount()):
                it = table.item(r, c)
                vals.append(it.text() if it else "")
            rows.append(vals)
        return rows

    def append_log(self, text):
        # 行数上限由 document().setMaximumBlockCount() 内部高效截断，这里只做追加
        self.log.append(text)
        self.log.moveCursor(QTextCursor.MoveOperation.End)

    def run_local(self, cmd, on_done=None):
        """本地子进程（cmd 为单条或多条命令）"""
        cmds = cmd if cmd and isinstance(cmd[0], list) else [cmd]
        if len(cmds) > 1:
            self.append_log(f"[本地] 共 {len(cmds)} 条命令")
        else:
            self.append_log("$ " + " ".join(cmds[0]))
        self.set_running(True)
        self.worker = Worker(cmds, BASE_DIR, self)
        self.worker.line.connect(self.append_log)
        self.worker.finished.connect(lambda rc: self.on_done(rc, on_done))
        self.worker.start()

    def run_remote(self, command, files_to_upload=None, on_done=None,
                   download_config=None):
        """SSH 远程执行"""
        host = self.edit_host.text().strip()
        user = self.edit_user.text().strip()
        pwd = self.edit_pass.text()
        port = self.spin_ssh_port.value()
        remote_dir = self.edit_remote_dir.text().strip()
        if not (host and user and pwd):
            QMessageBox.warning(self, "提示", "请填写远程主机/用户名/密码")
            return
        self.set_running(True)
        self.append_log(f"[SSH] 连接 {user}@{host}:{port} 目录 {remote_dir}")
        self.worker = SshWorker(host, port, user, pwd, remote_dir, command,
                                files_to_upload, download_config, self)
        self.worker.line.connect(self.append_log)
        self.worker.finished.connect(lambda rc: self.on_done(rc, on_done))
        self.worker.start()

    def on_done(self, rc, on_done=None):
        self.append_log(f"[完成] 退出码 {rc}")
        self.set_running(False)
        if on_done:
            on_done(rc)

    def set_running(self, running):
        self.btn_send.setEnabled(not running)
        self.btn_make.setEnabled(not running)
        self.btn_stop.setEnabled(running)
        if hasattr(self, "btn_upload_all"):
            self.btn_upload_all.setEnabled(not running)
        if hasattr(self, "btn_read_redis"):
            self.btn_read_redis.setEnabled(not running)
            self.btn_save_redis.setEnabled(not running)
        for b in ("btn_pick_dir", "btn_summary_refresh", "btn_summary_export"):
            if hasattr(self, b):
                getattr(self, b).setEnabled(not running)

    # ---------- 操作 ----------
    def on_make(self):
        names = self.selected_interfaces()
        if not names:
            QMessageBox.warning(self, "提示", "请至少勾选一个接口")
            return
        cmds = [[PYTHON, os.path.join(BASE_DIR, "make_excel.py"), "--interface", n]
                for n in names]
        self.run_local(cmds, on_done=lambda rc: self.update_excel_label())

    def selected_types(self):
        """返回选中的用例类型列表"""
        sel = []
        if self.chk_type_normal.isChecked():
            sel.append("normal")
        if self.chk_type_error.isChecked():
            sel.append("error")
        if self.chk_type_destroy.isChecked():
            sel.append("destroy")
        return sel

    def build_send_cmd(self, name):
        """构造 send_test.py 的命令行"""
        parts = ["python3", "send_test.py", "--interface", name,
                 "--workers", str(self.spin_workers.value()),
                 "--max", str(self.spin_max.value()),
                 "--wait", str(self.spin_wait.value())]
        types = self.selected_types()
        if types and len(types) < 3:
            parts.append("--type")
            parts.append(",".join(types))
        if self.chk_mock.isChecked():
            parts.append("--mock")
        else:
            parts.append("--no-mock")
        if self.chk_destroy_plugin.isChecked():
            parts.append("--destroy-via-plugin")
        if self.chk_quiet.isChecked():
            parts.append("--quiet")
        dm = self.combo_destroy_mode.currentData()
        if dm and dm != "type1":   # type1 是默认，不传参数
            parts.append("--destroy-mode")
            parts.append(dm)
        return " ".join(parts)

    def _save_ui_config(self):
        """把当前界面上的配置保存到 config.ini（基于已有配置增量更新，避免覆盖 Redis 等项）"""
        cfg = load_config()
        cfg.update({
            "host": self.edit_host.text().strip(),
            "port": str(self.spin_ssh_port.value()),
            "username": self.edit_user.text().strip(),
            "password": self.edit_pass.text(),
            "remote_dir": self.edit_remote_dir.text().strip(),
            "workers": str(self.spin_workers.value()),
            "max": str(self.spin_max.value()),
            "wait": str(self.spin_wait.value()),
            "mock": "1" if self.chk_mock.isChecked() else "0",
            "quiet": "1" if self.chk_quiet.isChecked() else "0",
            "destroy_via_plugin": "1" if self.chk_destroy_plugin.isChecked() else "0",
            "destroy_mode": self.combo_destroy_mode.currentData() or "type1",
            "remote": "1" if self.chk_remote.isChecked() else "0",
            "download": "1" if self.chk_download.isChecked() else "0",
            "box_redis": "1" if self.box_redis.isExpanded() else "0",
            "box_linux": "1" if self.box_linux.isExpanded() else "0",
            "box_out": "1" if self.box_out.isExpanded() else "0",
            "download_dir": self.edit_download_dir.text().strip() or DEFAULT_CONFIG["download_dir"],
        })
        # 分栏比例（界面已构建时才存在）
        if hasattr(self, "hsplit"):
            cfg["hsplit"] = ",".join(str(x) for x in self.hsplit.sizes())
        if hasattr(self, "vsplit"):
            cfg["vsplit"] = ",".join(str(x) for x in self.vsplit.sizes())
        save_config(cfg)

    def on_read_redis(self):
        """从远程拉取 DataHub.ini 的 [REDIS] 段显示到界面"""
        import paramiko
        try:
            c = paramiko.SSHClient()
            c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            c.connect(self.edit_host.text().strip(), port=self.spin_ssh_port.value(),
                      username=self.edit_user.text().strip(),
                      password=self.edit_pass.text(), timeout=15)
            sftp = c.open_sftp()
            path = f"{self.edit_remote_dir.text().strip()}/DataHub.ini"
            with sftp.open(path, "r") as fh:
                text = fh.read().decode("utf-8", "replace")
            sftp.close()
            c.close()
            vals = {}
            for line in text.splitlines():
                line = line.strip()
                for key in ("REDISHOST", "REDISPORT", "REDISPWD", "REDISSELECT"):
                    if line.startswith(key + "="):
                        vals[key] = line.split("=", 1)[1].strip()
            self.edit_r_host.setText(vals.get("REDISHOST", self.edit_r_host.text()))
            self.edit_r_port.setText(vals.get("REDISPORT", self.edit_r_port.text()))
            if vals.get("REDISPWD"):
                self.edit_r_pwd.setText(vals["REDISPWD"])
            try:
                self.spin_r_db.setValue(int(vals.get("REDISSELECT", "0")))
            except ValueError:
                pass
            self.append_log(f"[SSH] 已读取远程 Redis 配置: {vals}")
        except Exception as e:
            QMessageBox.warning(self, "提示", f"读取远程 Redis 配置失败: {e}")

    def on_save_redis(self):
        """把界面的 Redis 配置写入远程 DataHub.ini [REDIS] 段"""
        kv = {
            "REDISHOST": self.edit_r_host.text().strip(),
            "REDISPORT": self.edit_r_port.text().strip(),
            "REDISPWD": self.edit_r_pwd.text(),
            "REDISSELECT": str(self.spin_r_db.value()),
        }
        cfg_save = dict(load_config())
        cfg_save.update({"r_host": kv["REDISHOST"], "r_port": kv["REDISPORT"],
                         "r_pwd": kv["REDISPWD"], "r_db": kv["REDISSELECT"]})
        save_config(cfg_save)
        self._save_ui_config()
        worker = SshWorker(
            self.edit_host.text().strip(), self.spin_ssh_port.value(),
            self.edit_user.text().strip(), self.edit_pass.text(),
            self.edit_remote_dir.text().strip(), "echo 'Redis配置已更新'",
            ini_update={"kv": kv}, parent=self)
        worker.line.connect(self.append_log)
        worker.start()

    def on_upload_all(self):
        """批量上传所有接口数据(xlsx) + 接口定义(py)到远程"""
        if not self.chk_remote.isChecked():
            QMessageBox.warning(self, "提示", "请先勾选“启用远程执行”")
            return
        rd = self.edit_remote_dir.text().strip()
        files = []
        # 所有接口定义 py
        for f in sorted(os.listdir(INTERFACES_DIR)):
            if f.endswith(".py"):
                files.append((os.path.join(INTERFACES_DIR, f), f"{rd}/interfaces/{f}"))
        # 所有数据 xlsx
        if os.path.isdir(DATA_DIR):
            for f in sorted(os.listdir(DATA_DIR)):
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    files.append((os.path.join(DATA_DIR, f), f"{rd}/data/{f}"))
        # 脚本文件
        for f in ("send_test.py", "mock_datahub.py", "make_excel.py"):
            files.append((os.path.join(BASE_DIR, f), f"{rd}/{f}"))

        self._save_ui_config()
        self.set_running(True)
        self.append_log(f"[SSH] 批量上传 {len(files)} 个文件到 {rd}...")
        self.worker = SshWorker(self.edit_host.text().strip(), self.spin_ssh_port.value(),
                                self.edit_user.text().strip(), self.edit_pass.text(),
                                rd, "echo '上传完成'", files, self)
        self.worker.line.connect(self.append_log)
        self.worker.finished.connect(lambda rc: self.set_running(False))
        self.worker.start()

    def on_send(self):
        """批量发送：对勾选的每个接口依次执行（远程或本地）"""
        names = self.selected_interfaces()
        if not names:
            QMessageBox.warning(self, "提示", "请至少勾选一个接口")
            return
        self._save_ui_config()
        self._batch_names = names
        self._batch_idx = 0
        self._batch_stopped = False
        self._batch_start = time.time()
        # 批次参数快照：随汇总总表导出到"批次信息"sheet
        self._batch_params = {
            "时间": time.strftime("%Y-%m-%d %H:%M:%S"),
            "接口数": len(names),
            "接口": "、".join(names),
            "并发线程数(workers)": str(self.spin_workers.value()),
            "最多条数(max)": str(self.spin_max.value()),
            "等待回复秒数(wait)": str(self.spin_wait.value()),
            "模拟应答器(mock)": "开" if self.chk_mock.isChecked() else "关",
            "破坏数据走插件": "开" if self.chk_destroy_plugin.isChecked() else "关",
            "破坏类型": self.combo_destroy_mode.currentData() or "type1",
            "安静模式": "开" if self.chk_quiet.isChecked() else "关",
            "远程执行": "开" if self.chk_remote.isChecked() else "关",
            "远程主机": self.edit_host.text().strip(),
            "远程目录": self.edit_remote_dir.text().strip(),
        }
        self.set_running(True)
        self.tabs.setCurrentWidget(self.tab_log)      # 运行中看实时日志
        if len(names) > 1:
            self.append_log(f"\n[BATCH] 共 {len(names)} 个接口待发送: {', '.join(names)}")
        self._run_next_batch_item()

    def _download_config_for(self, name):
        """下载配置：stats JSON 始终下载(供汇总分析)；xlsx/运行日志仅在勾选“自动下载”时下载"""
        rd = self.edit_remote_dir.text().strip() + "/out"
        local_perf = self.edit_download_dir.text().strip() or \
                     os.path.join(BASE_DIR, "out", "performance")
        dirs = [
            {   # 汇总数据源：始终下载（小文件）
                "remote": rd + "/performance",
                "local": local_perf,
                "patterns": [f"{name}_*_stats.json"],
            },
        ]
        if self.chk_download.isChecked():
            dirs[0]["patterns"] += [f"{name}_*.xlsx"]
            # 运行日志：独立 out/logs 目录，仅勾选时下载
            dirs.append({
                "remote": rd + "/logs",
                "local": os.path.join(os.path.dirname(local_perf), "logs"),
                "patterns": [f"{name}_*.log"],
            })
        return {"dirs": dirs}

    def _run_next_batch_item(self):
        if self._batch_stopped:
            self.append_log("[BATCH] 已手动停止")
            self.set_running(False)
            return
        names = self._batch_names
        idx = self._batch_idx
        if idx >= len(names):
            self.append_log(f"[BATCH] 全部完成（共 {len(names)} 个接口）")
            self._auto_export = True
            self.tabs.setCurrentWidget(self.tab_summary)   # 跑完直接看汇总
            self._refresh_summary()   # 汇总分析始终刷新（stats JSON 始终下载）
            self.set_running(False)
            return
        name = names[idx]
        if len(names) > 1:
            self.append_log(f"[BATCH] ({idx + 1}/{len(names)}) 接口 {name} 开始...")

        cmd_str = self.build_send_cmd(name)
        local_dl_dir = self.edit_download_dir.text().strip() or \
                       os.path.join(BASE_DIR, "out", "performance")
        dl = self._download_config_for(name)

        if self.chk_remote.isChecked():
            self.worker = SshWorker(
                self.edit_host.text().strip(), self.spin_ssh_port.value(),
                self.edit_user.text().strip(), self.edit_pass.text(),
                self.edit_remote_dir.text().strip(), cmd_str,
                files_to_upload=self.remote_upload_files(name),
                download_config=dl, parent=self)
            self.worker.line.connect(self.append_log)
            self.worker.finished.connect(lambda rc: self.on_batch_item_done(rc))
            self.worker.start()
        else:
            # 本地模式（.so 在 Linux 时才有真实发送）
            args = cmd_str.split()[2:]   # 跳过 "python3 send_test.py"
            cmd = [PYTHON, os.path.join(BASE_DIR, "send_test.py")] + args
            self.worker = Worker(cmd, BASE_DIR, self)
            self.worker.line.connect(self.append_log)
            self.worker.finished.connect(lambda rc: self.on_batch_item_done(rc))
            self.worker.start()

    def on_batch_item_done(self, rc):
        name = self._batch_names[self._batch_idx]
        self.append_log(f"[BATCH] 接口 {name} 完成，退出码 {rc}")
        self._batch_idx += 1
        self._run_next_batch_item()

    def remote_upload_files(self, name):
        """需要上传到远程的文件列表：脚本 + 公共模块 + 接口定义 + 数据文件 + DataHub.ini"""
        rd = self.edit_remote_dir.text().strip()
        return [
            (os.path.join(BASE_DIR, "send_test.py"), f"{rd}/send_test.py"),
            (os.path.join(BASE_DIR, "mock_datahub.py"), f"{rd}/mock_datahub.py"),
            (os.path.join(BASE_DIR, "perf_stats.py"), f"{rd}/perf_stats.py"),
            (os.path.join(INTERFACES_DIR, "_common.py"), f"{rd}/interfaces/_common.py"),
            (os.path.join(INTERFACES_DIR, f"{name}.py"), f"{rd}/interfaces/{name}.py"),
            (os.path.join(DATA_DIR, f"{name}.xlsx"), f"{rd}/data/{name}.xlsx"),
        ]

    def on_stop(self):
        self._batch_stopped = True
        if self.worker:
            self.worker.stop()
            self.append_log("[STOP] 请求停止...")

    def closeEvent(self, event):
        self._save_ui_config()
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(3000)
        loader = getattr(self, "_summary_loader", None)
        if loader and loader.isRunning():
            loader.wait(2000)
        event.accept()


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
