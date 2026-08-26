# -*- coding: utf-8 -*-
"""
通用测试数据生成器：根据接口定义生成 Excel 测试数据
====================================================
用法：
  python make_excel.py --interface query       # 生成 data/query.xlsx
  python make_excel.py --interface acc_sign    # 生成 data/acc_sign.xlsx

生成的 Excel 统一放在 data/ 目录。
"""
import argparse
import importlib
import os
import sys
import traceback

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
INTERFACES_DIR = os.path.join(BASE_DIR, "interfaces")

# 公共列宽度 + 字段列宽度
META_W = {"case_no": 12, "case_type": 14, "case_desc": 36, "expected": 40}


def load_interface(name):
    sys.path.insert(0, INTERFACES_DIR)
    try:
        mod = importlib.import_module(name)
    except ImportError:
        avail = [f[:-3] for f in os.listdir(INTERFACES_DIR)
                 if f.endswith(".py") and not f.startswith(("__", "_"))]
        sys.exit(f"[FAIL] 接口 {name} 不存在。可用接口: {avail}")
    for attr in ("NAME", "HEADERS", "ROWS"):
        if not hasattr(mod, attr):
            sys.exit(f"[FAIL] 接口 {name} 缺少 {attr} 定义")
    return mod


def main():
    ap = argparse.ArgumentParser(description="生成接口测试数据 Excel")
    ap.add_argument("--interface", required=True, help="接口名，如 query / acc_sign")
    args = ap.parse_args()

    mod = load_interface(args.interface)
    os.makedirs(DATA_DIR, exist_ok=True)
    out = os.path.join(DATA_DIR, f"{mod.NAME}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = mod.NAME

    # 表头
    for col, (key, zh) in enumerate(mod.HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=f"{zh}\n({key})")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # 数据（非法字符转义，如控制字符）
    def _sanitize(v):
        if isinstance(v, str) and ILLEGAL_CHARACTERS_RE.search(v):
            return ILLEGAL_CHARACTERS_RE.sub("[_]", v)
        return v

    for r, row in enumerate(mod.ROWS, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=_sanitize(v))

    # 列宽：公共列用 META_W，其余按内容定
    for c, (key, zh) in enumerate(mod.HEADERS, 1):
        if key in META_W:
            w = META_W[key]
        else:
            # 按表头中文长度 + 数据里最大长度估
            maxlen = max(len(str(zh)), *[len(str(r[c - 1])) for r in mod.ROWS])
            w = min(max(12, maxlen + 4), 30)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # 原子保存，被占用则存 v2
    tmp = out + ".tmp"
    wb.save(tmp)
    try:
        os.replace(tmp, out)
    except PermissionError:
        out2 = os.path.join(DATA_DIR, f"{mod.NAME}_v2.xlsx")
        wb.save(out2)
        print(f"[WARN] {out} 正被占用(Excel 可能已打开)，已保存到 {out2}")
        return
    n = len(mod.ROWS)
    print(f"[OK] 已生成 {out}，共 {n} 条用例")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
