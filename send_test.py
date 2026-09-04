# -*- coding: utf-8 -*-
"""
通用多线程压测脚本（支持任意接口 + 正常/错误/破坏测试）
========================================================
流程：
  data/{接口}.xlsx 逐行读取测试数据
    -> 按接口定义构造报文(JSON)
    -> 多线程并发通过 libdatahub_trade_plug.so 的 SendMQ 发送
    -> 插件内部异步经 Redis 送数据中台，回复走回调返回

case_type 分流：
  normal  - 压测数据：走插件 SendMQ（性能测试）。账号四要素与真实账号完全匹配、
            字段完整合法，走完整业务主链路，统计结果才有意义。
  probe   - 兼容性探测：走插件 SendMQ，但结果不确定（缺字段/属性不匹配/边界取值），
            可能走错误路径或返回空，因此【不计入压测指标】，只用于功能确认。
            跑性能统计请用 --type normal 把 probe 排除掉。
  error   - 错误数据：走插件 SendMQ（看插件如何处理异常/是否拒绝）
  destroy - 破坏数据：默认【直接写 Redis】(绕过插件，--destroy-via-plugin 可改为走插件)
            破坏类型 destroy_mode 四类（两个维度组合：核心字段 × task 内容），
            对应中台处理链路的四道关卡：
              ①来源/路由(核心字段) → ②JSON解析 → ③协议分发(create/query键) → ④业务校验(字段值)
              type1 - 核心字段乱填 + 业务数据畸形（Excel 行的 fuzz 报文）—— 破坏①+④
              type2 - 核心字段乱填 + 业务数据正确（取 Excel 第一条 normal 报文）—— 破坏①
              type3 - 核心字段正常 + task 非法 JSON（截断/空/纯文本/二进制垃圾）—— 破坏②
              type4 - 核心字段正常 + task 合法 JSON 但非协议格式（空对象/数组/未知键）—— 破坏③
            （error 用例走插件、task 合法且协议格式，只破坏④业务校验）
            --destroy-mode 选 type1/type2/type3/type4/mixed；mixed=四类轮发；
            空则按 Excel 行级 destroy_mode 列，列缺失/为 mixed 时同样四类轮发

用法：
  python send_test.py --interface query            # 发送 data/query.xlsx
  python send_test.py --interface acc_sign --no-send   # 预览报文
  python send_test.py --interface query --workers 16 --max 1000
  python send_test.py --interface query --destroy-via-plugin
"""
import argparse
import asyncio
import ctypes
import importlib
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
INTERFACES_DIR = os.path.join(BASE_DIR, "interfaces")
sys.path.insert(0, BASE_DIR)
sys.path.insert(0, INTERFACES_DIR)

from mock_datahub import MockDataHub, RespClient
from perf_stats import PerfStats

# redis.asyncio 可用性（redis-py >= 4.2 提供）；
# 不可用时破坏测试退回串行逐条直写（每条新建连接，性能较低但功能一致）
try:
    import redis.asyncio as _aioredis
    HAS_AIOREDIS = True
except Exception:
    HAS_AIOREDIS = False

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("缺少 openpyxl，请先执行: pip install openpyxl")

# ---------- Redis 配置（默认 + 环境变量；真正以 DataHub.ini 为准）----------
REDIS_HOST = os.environ.get("REDISHOST", "192.168.1.137")
REDIS_PASSWORD = os.environ.get("REDISPWD", "QianLong@2026&")
REDIS_PORT = int(os.environ.get("REDISPORT", "6379"))
REDIS_SELECT = int(os.environ.get("REDISSELECT", "0"))

# 请求流名（插件/破坏测试都写入这里，数据中台从此消费）
REQ_STREAM = "DataHub_req_stream"

# 安静模式：不打印每条报文/回复，只输出关键信息（大并发压测时减少 GUI 日志量）
QUIET = False

# 安静模式下同类错误最多逐条打印的条数，超出后仅计数（避免大批量失败刷爆 GUI 日志）
_QUIET_ERR_LIMIT = 5
_quiet_err_count = {}


def _log_quiet_error(tag, msg):
    """错误日志：非安静模式全部打印；安静模式下同类最多打印前 _QUIET_ERR_LIMIT 条。

    成功明细请直接用 `if not QUIET: print(...)`，本函数只用于错误/告警。
    """
    n = _quiet_err_count.get(tag, 0) + 1
    _quiet_err_count[tag] = n
    if not QUIET or n <= _QUIET_ERR_LIMIT:
        print(f"  [{tag}] {msg}")
    elif n == _QUIET_ERR_LIMIT + 1:
        print(f"  [{tag}] 安静模式：后续同类错误仅计数，不再逐条打印")


class _Tee:
    """把 stdout 同时写到终端和运行日志文件（每行即时 flush 到文件）"""

    def __init__(self, stream, file):
        self.stream = stream
        self.file = file

    def write(self, data):
        self.stream.write(data)
        self.file.write(data)
        self.file.flush()

    def flush(self):
        try:
            self.stream.flush()
            self.file.flush()
        except Exception:
            pass

    def __getattr__(self, name):
        # 转发其余属性/方法给真实 stdout（isatty/encoding/fileno 等），
        # 避免第三方库在 tee 期间访问 stdout 其他成员时抛 AttributeError
        return getattr(self.stream, name)


def load_redis_cfg(cfg_path):
    """从 DataHub.ini 读取 REDIS 配置（插件实际上读这个文件，不是环境变量）"""
    global REDIS_HOST, REDIS_PASSWORD, REDIS_PORT, REDIS_SELECT
    ini = os.path.join(cfg_path, "DataHub.ini")
    if not os.path.exists(ini):
        return
    import configparser
    cp = configparser.ConfigParser()
    try:
        cp.read(ini, encoding="utf-8")
    except Exception:
        try:
            cp.read(ini)
        except Exception:
            return
    if cp.has_section("REDIS"):
        sec = cp["REDIS"]
        REDIS_HOST = sec.get("REDISHOST", REDIS_HOST)
        REDIS_PASSWORD = sec.get("REDISPWD", REDIS_PASSWORD)
        REDIS_PORT = int(sec.get("REDISPORT", REDIS_PORT))
        REDIS_SELECT = int(sec.get("REDISSELECT", REDIS_SELECT))
        print(f"[INFO] DataHub.ini REDIS配置: host={REDIS_HOST} port={REDIS_PORT} select={REDIS_SELECT}")


SO_CANDIDATES = [
    os.environ.get("SO_PATH", ""),
    os.path.join(BASE_DIR, "libdatahub_trade_plug.so"),
    os.path.join(BASE_DIR, "..", "lib", "libdatahub_trade_plug.so"),
    "/home/yangsh/so_test/libdatahub_trade_plug.so",
    "/home/liuyi/tmp/lib/libdatahub_trade_plug.so",
]


def find_so(explicit=""):
    if explicit and os.path.exists(explicit):
        return explicit
    for p in SO_CANDIDATES:
        if p and os.path.exists(p):
            return p
    return None


def load_interface(name):
    try:
        mod = importlib.import_module(name)
    except ImportError:
        avail = [f[:-3] for f in os.listdir(INTERFACES_DIR)
                 if f.endswith(".py") and not f.startswith(("__", "_"))]
        sys.exit(f"[FAIL] 接口 {name} 不存在。可用接口: {avail}")
    for attr in ("NAME", "HEADERS", "ROWS", "build_payload"):
        if not hasattr(mod, attr):
            sys.exit(f"[FAIL] 接口 {name} 缺少 {attr}")
    return mod


def load_cases(excel, max_cases):
    wb = load_workbook(excel, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"[FAIL] {excel} 为空")

    def _key(h):
        # 用 ASCII 限定，避免 \w 在 Unicode 模式下匹配到中文（如 "(逗号分隔)"）
        m = re.search(r"\(([A-Za-z_][A-Za-z0-9_]*)\)", str(h))
        return m.group(1) if m else str(h).strip()

    headers = [_key(h) for h in rows[0]]
    cases = []
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        rec = {headers[i]: ("" if raw[i] is None else raw[i]) for i in range(len(headers))}
        rec["_no"] = rec.get("case_no") or f"C{len(cases) + 1}"
        rec["_type"] = (rec.get("case_type") or "normal").strip().lower()
        cases.append(rec)
    return cases


# ==================== 插件客户端 ====================
_ReplyCb = ctypes.CFUNCTYPE(None, ctypes.c_int, ctypes.c_char_p, ctypes.c_char_p)


class DataHubClient:
    """封装 libdatahub_trade_plug.so 的 C API"""

    def __init__(self, so_path, unique="test", cfg_path=None, reply_flag=0):
        if cfg_path is None:
            cfg_path = BASE_DIR + "/"
        self.reply_flag = reply_flag
        self.lib = ctypes.CDLL(so_path)
        self.lib.CreateMQ.argtypes = [ctypes.c_char_p, ctypes.c_char_p, _ReplyCb]
        self.lib.CreateMQ.restype = ctypes.c_void_p
        self.lib.SendMQ.argtypes = [ctypes.c_void_p, ctypes.c_char_p, ctypes.c_char_p,
                                    ctypes.c_int, ctypes.c_char_p, ctypes.c_int]
        self.lib.SendMQ.restype = ctypes.c_int
        self.lib.get_version.restype = ctypes.c_char_p

        self._replies = []
        self._lock = threading.Lock()
        self._expect = 0                     # 期望回复条数（0=不限，由 wait_replies 设置）
        self._replied = threading.Event()    # 收齐期望条数时置位，用于提前结束等待

        def _on_msg(msg_id, data1, data2):
            s1 = data1.decode("utf-8", "replace") if data1 else ""
            s2 = data2.decode("utf-8", "replace") if data2 else ""
            with self._lock:
                self._replies.append((s2, s1))
                n = len(self._replies)
            # 收齐期望条数 -> 通知等待方立即返回，不必死等固定秒数
            if self._expect and n >= self._expect:
                self._replied.set()
            if not QUIET:
                print(f"  [reply] id={msg_id} req_id={s2} data={s1[:200]}")

        self._cb = _ReplyCb(_on_msg)
        self.mq = self.lib.CreateMQ(unique.encode(), cfg_path.encode(), self._cb)
        if not self.mq:
            raise RuntimeError("CreateMQ 返回空句柄")
        print(f"[OK] CreateMQ 成功，插件版本 {self.lib.get_version().decode()}，Redis={REDIS_HOST}")

    def _send_raw(self, payload, req_id, reply_flag, err):
        return self.lib.SendMQ(self.mq, payload.encode(), req_id.encode(),
                               reply_flag, err, 128)

    def wait_ready(self, timeout=15.0, probe_payload=None):
        if probe_payload is None:
            probe_payload = json.dumps(
                {"query": {"Account": {"AccountType": 1, "AccAtt": 6, "FAccount": "probe"}}})
        err = ctypes.create_string_buffer(128)
        start = time.time()
        while time.time() - start < timeout:
            ret = self._send_raw(probe_payload, "ready_probe", 0, err)
            if ret >= 0:
                print(f"[OK] 插件已 inited（{time.time() - start:.1f}s），可发送")
                return True
            time.sleep(0.5)
        print(f"[FAIL] {timeout}s 内插件未 inited")
        return False

    def send(self, payload, req_id, reply_flag=None):
        if reply_flag is None:
            reply_flag = self.reply_flag
        err = ctypes.create_string_buffer(128)
        ret = self._send_raw(payload, req_id, reply_flag, err)
        if ret < 0:
            _log_quiet_error("FAIL", f"send {req_id}: {err.value.decode('utf-8','replace')}")
        return ret

    @property
    def reply_count(self):
        """已收到的回复条数（回调线程并发写入，读取需加锁）"""
        with self._lock:
            return len(self._replies)

    def wait_replies(self, expected=0, timeout=0.0):
        """等待回复：收齐 expected 条立即返回，否则最多等 timeout 秒。

        expected<=0 表示不限条数，退化为固定等待 timeout 秒（保持旧行为）。
        返回实际收到的条数。
        """
        self._expect = max(0, expected)
        if self._expect and self.reply_count >= self._expect:
            return self.reply_count          # 发送期间就已收齐
        if timeout > 0:
            self._replied.wait(timeout)
        return self.reply_count


# ==================== 破坏测试：直接写 Redis ====================
# 按实测 DataHub_req_stream 的 6 字段格式写入，字段值与插件真实行为一致：
#   request_id / server_id / server_type / reply_req_stream / reply_reply_stream / task
# 破坏测试分四类（两个维度组合：核心字段 × task 内容）：
#   中台处理链路：①来源/路由(核心字段) → ②JSON解析 → ③协议分发 → ④业务校验
#   type1 - 核心字段乱填 + 业务数据畸形（Excel 行的 fuzz 报文）—— 破坏①+④
#   type2 - 核心字段乱填 + 业务数据正确（取 Excel 第一条 normal 报文）—— 破坏①
#   type3 - 核心字段正常 + task 非法 JSON —— 破坏②解析层（测解析容错不崩溃）
#   type4 - 核心字段正常 + task 合法 JSON 但非协议格式 —— 破坏③分发层（测未知结构容错）
#   （type3/4 核心字段必须正确，坏 task 才能穿透路由到达目标关卡）
_DESTROY_MODES = ("type1", "type2", "type3", "type4")
_BADFIELD_MODES = ("type1", "type2")   # 核心字段乱填的两种模式，共用 _TYPE2_TEMPLATES

# type3 模板池：task 为非法 JSON（破坏②解析层，中台解析必然失败，测解析容错不崩溃）
_BAD_JSON_TASKS = [
    "",                                     # 空串
    "{",                                    # 截断
    '{"create": ',                          # 键后截断
    '{"create": {"Account": {"FAccount"',   # 深处截断
    "not a json at all",                    # 纯文本
    "{{{{{{{{",                             # 括号垃圾
    '{"create": "\x00\x01\x02"}',           # 控制字符
    "abc\x00\x01\x02\xff\xfe",              # 二进制垃圾
    "\ufeff{\"create\":{}}",                # BOM 前缀（多数解析器直接失败）
    "{\r\t\n  ,,, }",                       # 非法语法
]

# type4 模板池：task 为合法 JSON 但不是协议格式（破坏③分发层，测未知结构的分发容错）
_NOT_PROTO_TASKS = [
    "{}",                                   # 空对象
    "[]",                                   # 顶层数组
    "null",                                 # 顶层 null
    "true",                                 # 顶层布尔
    "12345",                                # 顶层数字
    '"just a string"',                      # 顶层字符串
    '{"foo": "bar"}',                       # 未知顶层键
    '{"MsgType": 4}',                       # 只有内部协议键，缺业务体
    '{"create": "i am a string"}',          # create 不是对象
    '{"create": []}',                       # create 是数组
    '{"query": {"unknown_field": 1}}',      # 已知键但业务体字段未知
    '{"a": {"b": {"c": {"d": [1, 2, 3]}}}}',  # 深嵌套未知结构
]

# 核心字段乱填模板（type1/type2 共用，按 req_id 稳定哈希取模，保证可复现）：
#   {req_id} 会被替换为实际用例编号。每套组合覆盖一类破坏：
#   0 特殊符号+超大数字 / 1 类型错误 / 2 超长溢出 / 3 空值空白 / 4 控制字符 / 5 混合乱填
_TYPE2_TEMPLATES = [
    {  # 0 特殊符号包裹 + 超大数字 + 乱值
        "request_id": "$$${req_id}###",
        "server_id": "999999999999",
        "server_type": "XXX",
        "reply_req_stream": "no_such_stream",
        "reply_reply_stream": "garbage_reply",
    },
    {  # 1 类型错误：server_id 喂非数字（其余字段正常）
        "request_id": "$$${req_id}###",
        "server_id": "abc",
        "server_type": "WT",
        "reply_req_stream": "WT-abc",
        "reply_reply_stream": "WT-abc-reply",
    },
    {  # 2 超长溢出：各字段塞超长串
        "request_id": "{req_id}" + "X" * 200,
        "server_id": "123456789012345678901234567890",
        "server_type": "A" * 300,
        "reply_req_stream": "S" * 200,
        "reply_reply_stream": "R" * 200,
    },
    {  # 3 空值/空白
        "request_id": "",
        "server_id": "   ",
        "server_type": "",
        "reply_req_stream": " ",
        "reply_reply_stream": "",
    },
    {  # 4 控制字符/特殊符号
        "request_id": "WT\x00\x01-{req_id}",
        "server_id": "1e5",
        "server_type": "W\x02T",
        "reply_req_stream": "WT-\x03-0",
        "reply_reply_stream": "reply\x00garbage",
    },
    {  # 5 混合乱填：十六进制+乱值+不存在的流
        "request_id": "###{req_id}%%%",
        "server_id": "0x1F",
        "server_type": "sysadmin",
        "reply_req_stream": "wrong_stream_123",
        "reply_reply_stream": "???",
    },
]


def build_destroy_fields(payload, req_id, destroy_mode="type1", server_id="12345"):
    """构造破坏数据的 stream 字段。返回 (fields, tpl_idx)；核心字段正常时 tpl_idx=None。

    中台处理链路：①来源/路由(核心字段) → ②JSON解析 → ③协议分发 → ④业务校验
    type1(破坏①+④): 核心字段乱填（_TYPE2_TEMPLATES 按 req_id 稳定哈希取一套），
                     task 由调用方给定（业务畸形报文）
    type2(破坏①):   核心字段乱填，task 由调用方给定（业务正确报文）
    type3(破坏②):   核心字段正确（对齐插件真实值），task 为非法 JSON 模板
    type4(破坏③):   核心字段正确，task 为非协议 JSON 模板
                     （坏 task 见 _BAD_JSON_TASKS / _NOT_PROTO_TASKS，由调用方替换好）
    """
    if destroy_mode in _BADFIELD_MODES:
        # 乱填核心字段：按 req_id 稳定哈希取整套模板，保证可复现
        idx = sum(ord(ch) for ch in str(req_id)) % len(_TYPE2_TEMPLATES)
        tpl = _TYPE2_TEMPLATES[idx]
        fields = {
            k: v.format(req_id=req_id) if isinstance(v, str) else v
            for k, v in tpl.items()
        }
        fields["task"] = payload
    elif destroy_mode in _DESTROY_MODES:
        # type3/type4：核心字段正确，task 为坏 JSON 模板（调用方已替换好）
        idx = None
        fields = {
            "request_id": str(req_id),
            "server_id": server_id,
            "server_type": "WT",
            "reply_req_stream": f"WT-{server_id}",
            "reply_reply_stream": f"WT-{server_id}-reply",
            "task": payload,
        }
    else:
        raise ValueError(f"未知破坏类型: {destroy_mode}，可用: {_DESTROY_MODES}")
    return fields, idx


def destroy_write_redis(payload, req_id, destroy_mode="type1", server_id="12345"):
    """绕过插件，直接向 DataHub_req_stream 写入破坏数据（串行单条）。返回 True=成功。"""
    try:
        conn = RespClient(REDIS_HOST, REDIS_PORT, REDIS_PASSWORD, db=REDIS_SELECT)
        conn.connect()
        fields, idx = build_destroy_fields(payload, req_id, destroy_mode, server_id)
        # 打平成 field1 value1 field2 value2 ...
        flat = []
        for k, v in fields.items():
            flat.append(k)
            flat.append(v)
        conn.cmd("XADD", REQ_STREAM, "*", *flat)
        conn.close()
        if not QUIET:
            tag = f"#{idx}" if idx is not None else ""
            print(f"  [DESTROY/{destroy_mode}{tag}] 已直写 DataHub_req_stream: {req_id}"
                  f" server_id={fields.get('server_id', '')}")
        return True
    except Exception as e:
        _log_quiet_error("DESTROY", f"失败 {req_id}: {e}")
        return False


# ---------------------- 破坏测试批量直写（asyncio + pipeline） ----------------------
async def _destroy_send_async(items, stats, batch_size=500, concurrency=4):
    """asyncio 批量 XADD：pipeline 分批 + 并发控制。
    items: [(payload, req_id, destroy_mode, server_id), ...]
    返回成功条数。pipeline 无法测单条耗时，record_send 用批内平均延迟近似。"""
    client = _aioredis.Redis(
        host=REDIS_HOST, port=REDIS_PORT, password=REDIS_PASSWORD,
        db=REDIS_SELECT, decode_responses=True, socket_timeout=10)
    sem = asyncio.Semaphore(max(1, concurrency))

    async def send_batch(batch):
        async with sem:
            flat_list = []
            for payload, req_id, mode, server_id in batch:
                fields, _ = build_destroy_fields(payload, req_id, mode, server_id)
                flat = []
                for k, v in fields.items():
                    flat.append(k)
                    flat.append(v)
                flat_list.append(flat)
            t0 = time.perf_counter()
            pipe = client.pipeline(transaction=False)
            for flat in flat_list:
                pipe.execute_command("XADD", REQ_STREAM, "*", *flat)
            results = await pipe.execute(raise_on_error=False)
            us = (time.perf_counter() - t0) * 1e6 / max(1, len(batch))
            ok = 0
            for (payload, req_id, mode, _sid), flat, r in zip(batch, flat_list, results):
                ret = not isinstance(r, Exception)
                ok += ret
                if stats:
                    # 延迟为批内平均，kind="xadd" 与 SendMQ 延迟分开统计
                    stats.record_send(ret, len(payload.encode("utf-8")), us, kind="xadd")
                    # pipeline 结果确认才算成功写入（精确计数，与请求数对齐）
                    stats.record_redis_write(ret)
            return ok

    try:
        batches = [items[i:i + batch_size] for i in range(0, len(items), batch_size)]
        results = await asyncio.gather(*(send_batch(b) for b in batches),
                                       return_exceptions=True)
        ok = sum(r for r in results if isinstance(r, int))
        for r in results:
            if isinstance(r, Exception):
                _log_quiet_error("DESTROY", f"批量直写异常: {r}")
        return ok
    finally:
        await client.aclose()


def destroy_write_all(items, stats, batch_size=500, concurrency=4):
    """破坏测试批量直写总入口。返回 (成功条数, 耗时s, 途径描述)。
    优先 asyncio+pipeline（连接复用、批量网络往返）；无 redis 库时退回串行逐条。"""
    if HAS_AIOREDIS:
        t0 = time.time()
        ok = asyncio.run(_destroy_send_async(items, stats, batch_size, concurrency))
        return ok, time.time() - t0, f"asyncio+pipeline(批{batch_size}x并发{concurrency})"
    # 回退：串行逐条（旧逻辑，不依赖 redis 库）
    t0 = time.time()
    ok = 0
    for payload, req_id, mode, server_id in items:
        ts = time.perf_counter()
        ret = destroy_write_redis(payload, req_id, mode, server_id)
        us = (time.perf_counter() - ts) * 1e6
        if stats:
            stats.record_send(ret, len(payload.encode("utf-8")), us, kind="xadd")
            stats.record_redis_write(ret)
        ok += ret
    return ok, time.time() - t0, "串行逐条(未安装redis库)"


# ==================== 主流程 ====================
def main():
    ap = argparse.ArgumentParser(description="通用多线程压测")
    ap.add_argument("--interface", required=True, help="接口名，如 query / acc_sign")
    ap.add_argument("--excel", default="", help="Excel 路径（默认 data/{接口}.xlsx）")
    ap.add_argument("--so", default="", help=".so 路径，缺省自动查找")
    ap.add_argument("--workers", type=int, default=4, help="并发线程数")
    ap.add_argument("--max", type=int, default=0, help="最多处理多少条(0=全部)")
    ap.add_argument("--type", default="",
                    help="只发指定用例类型，逗号分隔，如 normal,error,destroy,probe"
                         "（空=全部；压测建议只传 normal）")
    ap.add_argument("--wait", type=float, default=3.0, help="发完后等待回复秒数")
    ap.add_argument("--reply", type=int, default=0, choices=[0, 1], help="reply_flag")
    ap.add_argument("--init-wait", type=float, default=5.0, help="等待插件 inited 最大秒数(兜底超时，正常2-3s即探测到)")
    ap.add_argument("--mock", action="store_true", default=True, help="启动模拟应答器")
    ap.add_argument("--no-mock", dest="mock", action="store_false", help="关闭模拟应答器")
    ap.add_argument("--destroy-via-plugin", action="store_true",
                    help="破坏数据也走插件 SendMQ（默认直接写 Redis）")
    ap.add_argument("--destroy-mode", default="",
                    help="破坏测试类型(type1/type2/type3/type4/mixed)："
                         "type1=乱填字段+业务畸形(破坏路由+业务校验)，"
                         "type2=乱填字段+业务正确(破坏路由)，"
                         "type3=正常字段+非法JSON(破坏解析层)，"
                         "type4=正常字段+非协议JSON(破坏分发层)；"
                         "mixed=四类轮发；空=按 Excel 行级 destroy_mode 列（缺列/mixed 同样四类轮发）")
    ap.add_argument("--destroy-server-id", default="12345",
                    help="type3/type4（核心字段正常）用的 server_id（默认 12345，与 mock 应答一致）")
    ap.add_argument("--no-stats", dest="stats", action="store_false", default=True,
                    help="关闭性能统计(默认开)")
    ap.add_argument("--stats-out", default="",
                    help="统计输出目录（默认 out/performance/）")
    ap.add_argument("--stats-interval", type=float, default=1.0,
                    help="统计采样间隔秒数")
    ap.add_argument("--stats-json", action="store_true", default=True,
                    help="额外保存 summary JSON（供 GUI 批量汇总，默认开）")
    ap.add_argument("--no-stats-json", dest="stats_json", action="store_false",
                    help="关闭 summary JSON 保存")
    ap.add_argument("--no-send", action="store_true", help="只生成报文不发送")
    ap.add_argument("--quiet", action="store_true",
                    help="安静模式：不打印每条报文/回复，只输出关键信息")
    args = ap.parse_args()
    global QUIET
    QUIET = args.quiet

    mod = load_interface(args.interface)

    # ---- 运行日志：tee stdout 到文件（排查问题用），统计明细最后追加 ----
    run_dir = os.path.join(BASE_DIR, "out", "logs")
    os.makedirs(run_dir, exist_ok=True)
    run_log = os.path.join(run_dir, f"{mod.NAME}_{time.strftime('%Y%m%d_%H%M%S')}.log")
    _orig_stdout = sys.stdout
    sys.stdout = _Tee(_orig_stdout, open(run_log, "w", encoding="utf-8"))

    excel = args.excel or os.path.join(DATA_DIR, f"{mod.NAME}.xlsx")
    cases = load_cases(excel, 0)   # 先读全部
    # 空数据必须在这里拦掉：否则下面 --max 循环扩量时 base 为空会陷入死循环
    if not cases:
        sys.exit(f"[FAIL] {excel} 中无有效用例（表头之后没有数据行）")
    # --type 过滤：只发指定用例类型（normal/error/destroy，可逗号分隔）
    if args.type:
        allowed = {t.strip().lower() for t in args.type.split(",") if t.strip()}
        cases = [c for c in cases if c["_type"] in allowed]
        if not cases:
            sys.exit(f"[FAIL] 类型 {args.type} 过滤后无用例。"
                     f"可用类型: normal/probe/error/destroy")
    # --max 循环：过滤后再循环到 N 条（压测需要）
    if args.max > len(cases):
        base = cases[:]
        n0 = len(cases)
        while len(cases) < args.max:
            for c in base:
                if len(cases) >= args.max:
                    break
                c2 = dict(c)
                c2["_no"] = f"{c2['_no']}_{len(cases) - n0 + 1}"
                cases.append(c2)
    elif 0 < args.max <= len(cases):
        cases = cases[:args.max]
    print(f"[INFO] 接口={mod.NAME} 从 {excel} 读取 {len(cases)} 条用例")
    from collections import Counter
    _dist = Counter(c["_type"] for c in cases)
    print(f"[INFO] 类型分布: {dict(_dist)}")
    if "probe" in _dist:
        print("[WARN] 本次含 probe 用例（兼容性探测，结果不确定），"
              "其耗时/成功率不计入性能指标；压测请加 --type normal\n")
    else:
        print()

    # 构造报文（按接口 + 行）
    payloads = []
    for c in cases:
        try:
            p = json.dumps(mod.build_payload(c), ensure_ascii=False)
        except Exception as e:
            _log_quiet_error("WARN", f"{c['_no']} 报文构造失败: {e}，改用原文案")
            p = str(c.get("case_desc"))
        payloads.append(p)
        if not QUIET:
            print(f"  {c['_no']} [{c['_type']}] {p[:160]}")

    so = find_so(args.so)
    load_redis_cfg(BASE_DIR + "/")

    if args.no_send or so is None:
        if so is None and not args.no_send:
            print(f"\n[WARN] 未找到 .so（查找: {SO_CANDIDATES}），已切换为报文预览模式")
        out = os.path.join(BASE_DIR, "out", f"{mod.NAME}_requests.jsonl")
        os.makedirs(os.path.dirname(out), exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            for p in payloads:
                f.write(p + "\n")
        print(f"\n[DONE] 报文预览模式：已保存到 {out}（共 {len(payloads)} 条）")
        sys.stdout.flush()
        sys.stdout = _orig_stdout
        return

    # ---------- 1) 先启动性能统计与模拟应答器（必须先于 CreateMQ）----------
    # 插件在 CreateMQ 后立即发布上线(id=-1)；若应答器此时未完成订阅，
    # 就只能等下一轮心跳（数秒），表现为 inited 前长时间"卡顿"。
    # 因此顺序必须是：订阅就绪 -> CreateMQ -> 立刻收到上线 -> 即刻应答。

    # 性能统计采样线程
    stats = None
    if args.stats:
        stats = PerfStats(redis_cfg={
            "host": REDIS_HOST, "port": REDIS_PORT, "password": REDIS_PASSWORD,
            "db": REDIS_SELECT, "stream": REQ_STREAM,
        })
        stats.start(interval=args.stats_interval)

    mock = None
    if args.mock:
        beat_channels = ["tradeserver_online"]
        if REDIS_SELECT > 0:
            beat_channels.append(f"tradeserver_online_{REDIS_SELECT}")
        print(f"[INFO] (先于 CreateMQ) 启动模拟数据中台应答器，订阅 {beat_channels}...")
        mock = MockDataHub(REDIS_HOST, REDIS_PORT, REDIS_PASSWORD, channels=beat_channels)
        mock.start()
        time.sleep(0.3)   # 给订阅握手指令留一点时间

    # ---------- 2) CreateMQ：插件上线后即刻被 mock 应答 ----------
    client = DataHubClient(so, unique=mod.NAME, reply_flag=args.reply)

    try:
        if not client.wait_ready(timeout=args.init_wait):
            raise SystemExit("插件未就绪，无法发送")

        # 分流：normal/error 走插件；destroy 直写 Redis（或走插件）
        plugin_cases = []
        destroy_cases = []
        for c, p in zip(cases, payloads):
            if c["_type"] == "destroy" and not args.destroy_via_plugin:
                destroy_cases.append((c, p))
            else:
                plugin_cases.append((c, p))

        # type2 需要"业务正确"的 task：优先取 Excel 第一条 normal 用例的报文
        good_payload = next((p for c, p in zip(cases, payloads)
                             if c["_type"] == "normal"), None)

        # 标记发送阶段开始（真实吞吐统计用）
        if stats:
            stats.mark_send_start()

        # 1) 插件发送（normal + error）
        if plugin_cases:
            def do_send(p, req_id):
                t0 = time.perf_counter()
                ret = client.send(p, req_id)
                us = (time.perf_counter() - t0) * 1e6
                if stats:
                    stats.record_send(ret >= 0, len(p.encode("utf-8")), us)
                return ret

            t0 = time.time()
            with ThreadPoolExecutor(max_workers=args.workers) as ex:
                futs = [ex.submit(do_send, p, f"{c['_no']}_{i}")
                        for i, (c, p) in enumerate(plugin_cases)]
                ok = sum(1 for f in futs if f.result() >= 0)
            dt = time.time() - t0
            print(f"\n[RESULT] 插件发送成功 {ok}/{len(plugin_cases)}，耗时 {dt:.3f}s，"
                  f"平均 {len(plugin_cases) / dt:.0f} 条/s")

        # 2) 破坏测试：直写 Redis（type1~type4 四类）
        if destroy_cases:
            # --destroy-mode: type1/type2/type3/type4/mixed；空=按 Excel 行级 destroy_mode 列
            # 行级列缺失或为 mixed 时，同样四类轮发（type1→type2→type3→type4→type1...）
            def resolve_mode(c, i):
                m = str(args.destroy_mode or c.get("destroy_mode") or "").strip().lower()
                if not m or m == "mixed":
                    return _DESTROY_MODES[i % len(_DESTROY_MODES)]
                if m not in _DESTROY_MODES:
                    sys.exit(f"[FAIL] 未知破坏类型 {m}，可用: {_DESTROY_MODES + ('mixed',)}")
                return m

            # type2 需要业务正确的报文
            if good_payload is None:
                print("[WARN] Excel 无 normal 用例，type2 将退用 destroy 行自身报文（内容可能仍畸形）")
                good_payload = payloads[0] if payloads else "{}"
            # type3/type4 的 task 与 Excel 行无关，直接从模板池轮换取
            mode_count = dict.fromkeys(_DESTROY_MODES, 0)
            bj_idx = np_idx = 0
            items = []
            for i, (c, p) in enumerate(destroy_cases):
                mode = resolve_mode(c, i)
                task = p
                if mode == "type2":
                    task = good_payload
                elif mode == "type3":
                    task = _BAD_JSON_TASKS[bj_idx % len(_BAD_JSON_TASKS)]
                    bj_idx += 1
                elif mode == "type4":
                    task = _NOT_PROTO_TASKS[np_idx % len(_NOT_PROTO_TASKS)]
                    np_idx += 1
                mode_count[mode] += 1
                items.append((task, c["_no"], mode, args.destroy_server_id))
            dist = "，".join(f"{m}: {mode_count[m]}" for m in _DESTROY_MODES)
            print(f"\n[DESTROY] 破坏测试共 {len(destroy_cases)} 条（{dist}），直写 Redis...")
            ok, dt, via = destroy_write_all(items, stats,
                                            concurrency=max(2, min(args.workers, 8)))
            failed = len(destroy_cases) - ok
            print(f"[DESTROY] 直写成功 {ok}/{len(destroy_cases)}，耗时 {dt:.3f}s，{via}"
                  + (f"，失败 {failed} 条" if failed else ""))

        # 标记发送阶段结束
        if stats:
            stats.mark_send_end()

        # 收齐即停：期望条数 = 走插件的用例数（destroy 直写 Redis，不产生插件回复）
        expect = len(plugin_cases)
        if expect:
            print(f"[INFO] 等待回复（最多 {args.wait}s，收齐 {expect} 条即停）...")
            got = client.wait_replies(expect, args.wait)
            if got >= expect:
                print(f"[RESULT] 收到回复数: {got}/{expect}（已收齐，提前结束等待）")
            else:
                print(f"[RESULT] 收到回复数: {got}/{expect}（超时未收齐，缺 {expect - got} 条）")
        else:
            print(f"[INFO] 无走插件的用例，等待 {args.wait}s 收集残留回复...")
            got = client.wait_replies(0, args.wait)
            print(f"[RESULT] 收到回复数: {got}")
        with client._lock:
            sample = client._replies[:10]
        for rid, data in sample:
            print(f"    req_id={rid} -> {data[:200]}")

    finally:
        if mock:
            mock.stop()
        # 汇总数据 + 运行日志收尾
        if stats:
            try:
                stats.stop()
                out_dir = args.stats_out or os.path.join(BASE_DIR, "out", "performance")
                os.makedirs(out_dir, exist_ok=True)
                ts = time.strftime("%Y%m%d_%H%M%S")
                # summary JSON（供 GUI 批量汇总）
                if args.stats_json:
                    try:
                        json_path = os.path.join(out_dir, f"{mod.NAME}_{ts}_stats.json")
                        with open(json_path, "w", encoding="utf-8") as f:
                            json.dump({"interface": mod.NAME, "ts": ts,
                                       "summary": stats.summary()},
                                      f, ensure_ascii=False, indent=2)
                        print(f"[STATS] JSON 汇总已保存: {json_path}")
                    except Exception as e:
                        print(f"[WARN] JSON 汇总保存失败: {e}")
                print("[STATS] 汇总:")
                for k, v in stats.summary().items():
                    print(f"  {k}: {v}")
                # 按秒明细追加到运行日志
                with open(run_log, "a", encoding="utf-8") as f:
                    f.write(stats.detail_text())
                    f.write("\n")
            except Exception as e:
                print(f"[WARN] 统计输出失败: {e}")
                import traceback
                traceback.print_exc()
        print(f"[INFO] 运行日志已保存: {run_log}")
        print("[INFO] 插件有后台线程，直接强制退出（跳过 DestroyMQ）")
        sys.stdout.flush()
        sys.stdout = _orig_stdout
        os._exit(0)


if __name__ == "__main__":
    main()
